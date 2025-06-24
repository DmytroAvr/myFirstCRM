# oids/utils.py
import openpyxl
import datetime
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

def export_to_excel(queryset, columns, filename='export.xlsx', include_row_numbers=False):
    """
	:param queryset: QuerySet з даними для експорту.
    :param columns: Словник, де ключ - це рядок доступу до поля (напр. 'unit__code'), 
                    а значення - це назва стовпця в Excel.
    :param filename: Ім'я файлу для завантаження.
    :param include_row_numbers: Якщо True, додає стовпець '№' з нумерацією рядків.

    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Data'

    # --- 1. СТВОРЮЄМО СТИЛІ ---
    header_font = Font(name='Times New Roman', size=14, bold=True)
    data_font = Font(name='Times New Roman', size=14, bold=False)
    # Стиль для переносу тексту (висота рядка налаштується автоматично)
    wrap_alignment_header = Alignment(wrap_text=True, vertical='center', horizontal='center')
    wrap_alignment_data = Alignment(wrap_text=True, vertical='center', horizontal='left')

    # --- Записуємо заголовки ---
    start_col = 1
    if include_row_numbers:
        cell = sheet.cell(row=1, column=1)
        cell.value = '№'
        cell.font = header_font
        cell.alignment = wrap_alignment_header
        start_col = 2

    for col_num, column_title in enumerate(columns.values(), start_col):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = wrap_alignment_header

    # --- Записуємо дані ---
    for row_num, obj in enumerate(queryset, 2):
        row_in_list = row_num - 1
        
        if include_row_numbers:
            cell = sheet.cell(row=row_num, column=1)
            cell.value = f"{row_in_list}."
            cell.font = data_font
            cell.alignment = wrap_alignment_data

        for col_num, field_name in enumerate(columns.keys(), start_col):
            cell = sheet.cell(row=row_num, column=col_num)
            
            value = obj
            try:
                for part in field_name.split('__'):
                    attr = getattr(value, part)
                    if callable(attr):
                        value = attr()
                    else:
                        value = attr
            except (AttributeError, TypeError):
                value = None
            
			# --- ЛОГІКА ФОРМАТУВАННЯ ДАТ ---
            if isinstance(value, (datetime.date, datetime.datetime)):
                cell.value = value
                cell.number_format = 'dd.mm.yyyy'
            else:
                cell.value = str(value) if value is not None else ""

            cell.value = str(value) if value is not None else ""
            cell.font = data_font
            cell.alignment = wrap_alignment_data

    # --- 2. НАЛАШТУВАННЯ АВТОМАТИЧНОЇ ШИРИНИ СТОВПЦІВ ---
    for column_cells in sheet.columns:
        # +2 додає невеликий відступ для кращої читабельності
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 5)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # --- Налаштування друку (як і раніше) ---
    max_row = sheet.max_row
    max_col_letter = get_column_letter(sheet.max_column)
    if max_row > 0 and sheet.max_column > 0:
        sheet.print_area = f'A1:{max_col_letter}{max_row}'
    
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    workbook.save(response)
    return response