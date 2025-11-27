# :\myFirstCRM\oids\views.py
from django import forms 
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.management.base import BaseCommand
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.db.models import Q, Max, Prefetch, Count, OuterRef, Subquery
from django.db import  transaction
from django.utils import timezone
import datetime
from django.utils.safestring import mark_safe
from .utils import export_to_excel



from .models import (OIDTypeChoices, OIDStatusChoices, SecLevelChoices, WorkRequestStatusChoices, WorkTypeChoices, 
    DocumentReviewResultChoices, AttestationRegistrationStatusChoices, PeminSubTypeChoices, DocumentProcessingStatusChoices, add_working_days
)
from .models import (Unit, UnitGroup, OID, DskEot, OIDStatusChange, TerritorialManagement, 
	Document, DocumentType, WorkRequest, WorkRequestItem, Trip, TripResultForUnit, 
    Person, TechnicalTask, AttestationRegistration, AttestationResponse, WorkCompletionRegistration,
    ProcessTemplate, OIDProcess, OIDProcessStepInstance,
    Declaration, DeclarationRegistration,
)


from .forms import ( TripForm, TripResultSendForm, DocumentProcessingMainForm, DocumentItemFormSet, DocumentForm, 
    OIDCreateForm, OIDStatusUpdateForm, 
	WorkRequestForm, WorkRequestItemFormSet, TripResultForUnitForm,
    TechnicalTaskCreateForm, TechnicalTaskProcessForm,
	AttestationRegistrationSendForm, AttestationResponseMainForm, AttestationActUpdateFormSet,
    WorkCompletionSendForm, WorkCompletionResponseForm,    
	AzrUpdateForm, AzrItemFormSet, AzrSubmissionForm, AzrUpdateFormSet, 
    DeclarationSubmissionForm, DeclarationItemFormSet, DeclarationResponseForm, DeclarationUpdateFormSet
)
from .forms_process import ( DeclarationProcessStartForm, SendForRegistrationForm    
)
from .forms_filters import (
    WorkRequestFilterForm, OIDFilterForm, TechnicalTaskFilterForm, WorkRequestItemProcessingFilterForm, DocumentFilterForm,
    TechnicalTaskFilterForm,
    AttestationRegistrationFilterForm, AttestationResponseFilterForm,
    RegisteredActsFilterForm, AzrDocumentFilterForm,
    DeclarationFilterForm
)

# from .utils import add_working_days # Або перемістіть add_working_days сюди

# def add_working_days(start_date, days_to_add): ... (якщо не в utils.py)

def get_paginated_page(queryset, request, items_per_page=100):
    paginator = Paginator(queryset, items_per_page)
    page_number = request.GET.get('page')
    return paginator.get_page(page_number)
 
def ajax_get_oid_current_status(request):
    oid_id_str = request.GET.get('oid_id')
    if oid_id_str and oid_id_str.isdigit():
        oid_id = int(oid_id_str)
        try:
            oid_instance = OID.objects.get(pk=oid_id)
            return JsonResponse({
                'status': 'success',
                'current_status_display': oid_instance.get_status_display(),
                'current_status_value': oid_instance.status,
                'oid_cipher': oid_instance.cipher,
                'oid_type_display': oid_instance.get_oid_type_display()
            })
        except OID.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'ОІД не знайдено'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Необхідно ID ОІДа'}, status=400)

 # changede by gemeni. перейшли від передачі словників до передачі повних екземплярів моделі OID у функцію get_last_document_expiration_date. Це важливо для коректної роботи сервера, незалежно від фронтенд-фільтрації.

def ajax_load_oids_for_unit_categorized(request):
    unit_id_str = request.GET.get('unit_id')
    data = {
        'creating': [],
        'active': [],
        'cancelled': []
    }

    if unit_id_str:
        try:
            unit_id = int(unit_id_str)
            # Замість .values(), отримуємо повні об'єкти OID, щоб передавати їх у helper
            oids_for_unit_qs = OID.objects.filter(unit__id=unit_id)\
                                        .select_related('unit', 'unit__territorial_management')\
                                        .order_by('cipher')

            for oid_instance in oids_for_unit_qs: # Тепер це повний екземпляр OID
                oid_item = {
                    'id': oid_instance.id,
                    'cipher': oid_instance.cipher,
                    'full_name': oid_instance.full_name or oid_instance.unit.city,
                    'oid_type_display': oid_instance.get_oid_type_display(), # Використовуємо метод моделі
                    'status_display': oid_instance.get_status_display(),   # Використовуємо метод моделі
                    'detail_url': reverse('oids:oid_detail_view_name', args=[oid_instance.id])
                }

                if oid_instance.status in [OIDStatusChoices.NEW, OIDStatusChoices.RECEIVED_REQUEST, OIDStatusChoices.RECEIVED_REQUEST_ATTESTATION, OIDStatusChoices.RECEIVED_TZ]:
                    data['creating'].append(oid_item)
                # elif oid_instance.status == OIDStatusChoices.ACTIVE:
                elif oid_instance.status in [OIDStatusChoices.ACTIVE, OIDStatusChoices.RECEIVED_REQUEST_IK, OIDStatusChoices.RECEIVED_REQUEST_PLAND_ATTESTATION,]:
                    # Тепер передаємо повний екземпляр oid_instance
                    oid_item['ik_expiration_date'] = get_last_document_expiration_date(oid_instance, 'Висновок')
                    oid_item['attestation_expiration_date'] = get_last_document_expiration_date(oid_instance, 'Акт атестації')
                    oid_item['prescription_expiration_date'] = get_last_document_expiration_date(oid_instance, 'Припис')
                    data['active'].append(oid_item)
                elif oid_instance.status in [OIDStatusChoices.CANCELED, OIDStatusChoices.TERMINATED, OIDStatusChoices.INACTIVE]:
                    data['cancelled'].append(oid_item)
        
        except ValueError:
            return JsonResponse({'error': 'Невірний ID військової частини'}, status=400)
        except Exception as e:
            # Виводимо помилку в консоль Django для дебагу
            print(f"SERVER ERROR in ajax_load_oids_for_unit_categorized: {type(e).__name__} - {e}")
            import traceback
            traceback.print_exc() # Друкує повний трейсбек
            return JsonResponse({'error': f'Серверна помилка: {type(e).__name__}'}, status=500)
            
    return JsonResponse(data)
# Ваш ajax_load_oids_for_unit (якщо потрібен окремо для простого списку ОІДів, наприклад, для форм)
 
def ajax_load_oids_for_multiple_units(request):
    unit_ids_str = request.GET.getlist('unit_ids[]') # Отримуємо список ID як рядки
    # Або якщо JS надсилає як 'unit_ids' через кому: request.GET.get('unit_ids', '').split(',')
    
    oids_data = []
    unit_ids = []
    for uid_str in unit_ids_str:
        if uid_str.isdigit():
            unit_ids.append(int(uid_str))

    if unit_ids:
        try:
            # Отримуємо ОІДи, що належать до будь-якої з обраних ВЧ
            oids_queryset = OID.objects.filter(
                unit__id__in=unit_ids,
                # Можна додати фільтр за статусом ОІД, наприклад:
                # status__in=[OIDStatusChoices.ACTIVE, OIDStatusChoices.NEW, OIDStatusChoices.RECEIVED_REQUEST, OIDStatusChoices.RECEIVED_TZ]
            ).select_related('unit').order_by('unit__code', 'cipher').distinct()
            
            for oid in oids_queryset:
                oids_data.append({
                    'id': oid.id,
                    'cipher': oid.cipher, 
                    'full_name': oid.full_name or "",
                    'unit_code': oid.unit.code, # Додаємо код ВЧ для кращого відображення
                    'unit_city': oid.unit.city,
					'oid_type_display': oid.get_oid_type_display(),
                    'pemin_sub_type': oid.get_pemin_sub_type_display()
                })
        except ValueError:
            return JsonResponse({'error': 'Невірні ID військових частин'}, status=400)
            
    return JsonResponse(oids_data, safe=False)
 
def ajax_load_work_request_items_for_oid(request):
    oid_id_str = request.GET.get('oid_id')
    items_data = []
    if oid_id_str and oid_id_str.isdigit():
        oid_id = int(oid_id_str)
        try:
            # Отримуємо елементи заявок для цього ОІД, можливо, з певними статусами
            items_queryset = WorkRequestItem.objects.filter(
                oid_id=oid_id,
                # status__in=[WorkRequestStatusChoices.PENDING, WorkRequestStatusChoices.IN_PROGRESS] # Приклад фільтрації
            ).select_related('request').order_by('-request__incoming_date')  
            for item in items_queryset:
                items_data.append({
                    'id': item.id,
                    'text': f"Заявка №{item.request.incoming_number} ({item.request.incoming_date.strftime('%d.%m.%Y')}) - {item.get_work_type_display()}",
                    'work_type': item.work_type # Передаємо work_type для подальшої фільтрації DocumentType
                })
        except ValueError:
            return JsonResponse({'error': 'Невірний ID ОІД'}, status=400)    
    return JsonResponse(items_data, safe=False)

def ajax_load_oids_for_unit(request):
    unit_id_str = request.GET.get('unit_id')
    oids_data = []
    if unit_id_str and unit_id_str.isdigit():
        unit_id = int(unit_id_str)
        try:
            # Отримуємо всі ОІДи для вказаної ВЧ.
            # Можна додати додаткові фільтри, якщо потрібно (наприклад, тільки активні ОІДи)
            # OID.objects.filter(unit_id=unit_id, status=OIDStatusChoices.ACTIVE).order_by('cipher')
            oids_queryset = OID.objects.filter(unit_id=unit_id).order_by('cipher')
            
            for oid in oids_queryset:
                oids_data.append({
                    'id': oid.id,
                    'cipher': oid.cipher, 
                    'full_name': oid.full_name or "", # Повертаємо порожній рядок, якщо full_name None
					'oid_type': oid.oid_type, 
					'status': oid.status, 
				    # Додайте інші поля, якщо вони потрібні для відображення в TomSelect у JS:
                    # наприклад, 'unit_code': oid.unit.code (якщо unit вже завантажено через select_related у запиті)
                })
        except ValueError: # На випадок, якщо unit_id не є валідним числом (хоча isdigit вже перевіряє)
            return JsonResponse({'error': 'Невірний ID військової частини'}, status=400)
        # except Unit.DoesNotExist: # Якщо ВЧ з таким ID не існує
            # return JsonResponse({'error': 'Військова частина не знайдена'}, status=404) 
            # Зазвичай, якщо queryset порожній, повернеться порожній список, що є нормальним.
    
    return JsonResponse(oids_data, safe=False) # safe=False, оскільки ми повертаємо список

def ajax_load_document_types_for_oid_and_work(request):
    oid_id_str = request.GET.get('oid_id')
    work_type_from_wri = request.GET.get('work_type') # Тип робіт з обраного WorkRequestItem
    # Або можна передавати work_request_item_id і отримувати work_type з нього на сервері

    doc_types_data = []
    
    if oid_id_str and oid_id_str.isdigit():
        oid_id = int(oid_id_str)
        try:
            oid_instance = OID.objects.get(pk=oid_id)
            oid_actual_type = oid_instance.oid_type # Тип самого ОІД (МОВНА/ПЕМІН)

            # Фільтруємо DocumentType
            # 1. За типом ОІД (МОВНА/ПЕМІН) - має збігатися або бути "спільним" (припустимо, спільний має порожній oid_type у DocumentType)
            # 2. За типом робіт (Атестація/ІК) - має збігатися з work_type_from_wri або бути "спільним"
            
            q_filters = Q(oid_type=oid_actual_type) | Q(oid_type='') # Або порожній рядок, або спеціальне значення для "спільного"
            
            if work_type_from_wri: # Якщо тип робіт визначено (наприклад, з обраного WorkRequestItem)
                q_filters &= (Q(work_type=work_type_from_wri) | Q(work_type=''))
            # else:
                # Якщо тип робіт не визначено, можливо, показувати всі типи документів, 
                # що підходять для типу ОІД, або вимагати вибору типу робіт окремо.
                # Поки що, якщо work_type_from_wri не передано, фільтр по work_type не застосовується (або показує тільки спільні для роботи).
                # Для більшої точності, якщо work_type не передано, можливо, не варто повертати нічого або тільки дуже загальні типи.
                # Або додати окреме поле "Тип робіт" на форму, якщо WorkRequestItem не обрано.
                # Поточна логіка: якщо work_type_from_wri не надано, фільтр по work_type буде широким.

            doc_types_queryset = DocumentType.objects.filter(q_filters).order_by('name')
            
            for dt in doc_types_queryset:
                doc_types_data.append({
                    'id': dt.id,
                    # Відображаємо всю інформацію для ясності у випадаючому списку
                    'text': f"{dt.name} (для ОІД: {dt.get_oid_type_display()}, для робіт: {dt.get_work_type_display()})"
                })

        except OID.DoesNotExist:
            return JsonResponse({'error': 'ОІД не знайдено'}, status=404)
        except Exception as e:
            print(f"Error in ajax_load_document_types: {e}")
            return JsonResponse({'error': str(e)}, status=500)
            
    return JsonResponse(doc_types_data, safe=False)
# Цей view НЕ використовується для оновлення списків ОІД на головній панелі в цьому сценарії

# oids/urls.py - не забудьте додати:
# path('ajax/load-oids-for-unit/', views.ajax_load_oids_for_unit, name='ajax_load_oids_for_unit'),

def ajax_load_work_requests_for_oids(request):
    oid_ids_str = request.GET.getlist('oid_ids[]') # Отримуємо список ID ОІДів

    work_requests_data = []
    oid_ids = []
    for oid_str in oid_ids_str:
        if oid_str.isdigit():
            oid_ids.append(int(oid_str))

    if oid_ids:
        try:
            # Знаходимо заявки, які мають WorkRequestItem, пов'язаний з будь-яким із переданих ОІДів
            # і мають відповідний статус
            relevant_requests = WorkRequest.objects.filter(
                items__oid__id__in=oid_ids,
                status__in=[WorkRequestStatusChoices.PENDING, WorkRequestStatusChoices.IN_PROGRESS]
            ).select_related('unit').distinct().order_by('-incoming_date')
            # distinct() важливий, якщо одна заявка має кілька ОІДів з переданого списку

            for wr in relevant_requests:
                work_requests_data.append({
                    'id': wr.id,
                    'text': str(wr)
					# 'text': f"заявка вх.№ {wr.incoming_number} від {wr.incoming_date.strftime('%d.%m.%Y')} (ВЧ: {wr.unit.code}) - {wr.get_status_display()}"
                })
        except Exception as e:
            # Обробка можливих помилок, наприклад, якщо OID.DoesNotExist (малоймовірно при filter id__in)
            print(f"Error in ajax_load_work_requests_for_oids: {e}") # Для логування
            return JsonResponse({'error': str(e)}, status=500)
            
    return JsonResponse(work_requests_data, safe=False)

def ajax_load_attestation_acts_for_oid(request):
    oid_id_str = request.GET.get('oid_id')
    acts_data = []

    if oid_id_str and oid_id_str.isdigit():
        oid_id = int(oid_id_str)
        try:
            # Знаходимо тип документа "Акт атестації"
            # Краще мати більш надійний спосіб ідентифікації типу, наприклад, по slug або константі
            attestation_act_doc_type = DocumentType.objects.filter(name__icontains="Акт атестації").first()
            
            
            if attestation_act_doc_type:
                # Обираємо документи типу "Акт атестації" для даного ОІД,
                # які ще не були відправлені на реєстрацію
                acts_queryset = Document.objects.filter(
                    oid_id=oid_id,
                    document_type=attestation_act_doc_type,
                    attestation_registration_sent__isnull=True # Ключовий фільтр
                ).select_related('oid__unit').order_by('-work_date', '-doc_process_date')
                
                for doc in acts_queryset:
                    acts_data.append({
                        'id': doc.id,
                        'text': f"Акт №{doc.document_number} від {doc.work_date.strftime('%d.%m.%Y')} (ОІД: {doc.oid.cipher})"
                        # Можна додати більше інформації, якщо потрібно для відображення
                    })
            else:
                print("AJAX_LOAD_ACTS: DocumentType 'Акт атестації'(name__icontains=Акт атестації) not found.")
        except ValueError:
            return JsonResponse({'error': 'Невірний ID ОІД'}, status=400)
        except Exception as e:
            print(f"Error in ajax_load_attestation_acts_for_oid: {e}")
            return JsonResponse({'error': 'Серверна помилка при завантаженні актів'}, status=500)
            
    return JsonResponse(acts_data, safe=False)

def ajax_load_attestation_acts_for_multiple_oids(request):
    oid_ids_str = request.GET.getlist('oid_ids[]')
    acts_data = []
    
    oid_ids = []
    for oid_str in oid_ids_str:
        if oid_str.isdigit():
            oid_ids.append(int(oid_str))

    if oid_ids:
        try:
            # Ваш спосіб ідентифікації типу документа "Акт атестації"
            attestation_act_doc_type = DocumentType.objects.filter(name__icontains="Акт атестації").first() 
            # Переконайтеся, що цей фільтр надійний, або використовуйте ID/slug типу документа

            if attestation_act_doc_type:
                acts_queryset = Document.objects.filter(
                    oid_id__in=oid_ids, # Акти, що належать до будь-якого з обраних ОІДів
                    document_type=attestation_act_doc_type,
                    attestation_registration_sent__isnull=True # Ще не відправлені
                ).select_related('oid__unit').order_by('oid__unit__code', 'oid__cipher', '-work_date')
                
                for doc in acts_queryset:
                    acts_data.append({
                        'id': doc.id,
                        'text': f"Акт №{doc.document_number} від {doc.work_date.strftime('%d.%m.%Y')} (ОІД: {doc.oid.cipher}, ВЧ: {doc.oid.unit.code})"
                    })
            else:
                print("AJAX_LOAD_ACTS_MULTIPLE_OIDS: DocumentType 'Акт атестації' (name__icontains=Акт атестації) not found.")
        except Exception as e:
            print(f"Error in ajax_load_attestation_acts_for_multiple_oids: {e}")
            return JsonResponse({'error': 'Серверна помилка при завантаженні актів'}, status=500)
            
    return JsonResponse(acts_data, safe=False)

# trip results
def ajax_load_units_for_trip(request):
    trip_id_str = request.GET.get('trip_id')
    units_data = []
    if trip_id_str and trip_id_str.isdigit():
        trip_id = int(trip_id_str)
        try:
            trip = Trip.objects.get(pk=trip_id)
            # Беремо ВЧ, що були задіяні у відрядженні
            for unit in trip.units.all().order_by('code'):
                units_data.append({
                    'id': unit.id, 
                    'text': f"{unit.code} - {unit.city or unit.name}"
                    })
        except Trip.DoesNotExist:
            pass # Повернемо порожній список
        except Exception as e:
            print(f"Error in ajax_load_units_for_trip: {e}")
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse(units_data, safe=False)

def ajax_load_oids_for_trip_units(request):
    trip_id_str = request.GET.get('trip_id')
    unit_ids_str = request.GET.getlist('unit_ids[]')
    oids_data = []

    unit_ids = [int(uid) for uid in unit_ids_str if uid.isdigit()]

    if trip_id_str and trip_id_str.isdigit() and unit_ids:
        trip_id = int(trip_id_str)
        try:
            trip = Trip.objects.get(pk=trip_id)
            # ОІДи, які були у відрядженні ТА належать до обраних ВЧ
            oids_queryset = trip.oids.filter(unit_id__in=unit_ids).select_related('unit').distinct().order_by('unit__code', 'cipher')
            for oid in oids_queryset:
                oids_data.append({
                    'id': oid.id, 
                    'text': f"ВЧ: {oid.unit.code} ОІД: {oid.cipher} - {oid.full_name or oid.unit.city}",
                    'oid_type': oid.oid_type # Передаємо тип ОІД для подальшої фільтрації документів
                })
        except Trip.DoesNotExist:
            pass
        except Exception as e:
            print(f"Error in ajax_load_oids_for_trip_units: {e}")
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse(oids_data, safe=False)

def ajax_load_documents_for_trip_oids(request):
    trip_id_str = request.GET.get('trip_id')
    oid_ids_str = request.GET.getlist('oid_ids[]')
    documents_data = []

    oid_ids = [int(oid_str) for oid_str in oid_ids_str if oid_str.isdigit()]

    if trip_id_str and trip_id_str.isdigit() and oid_ids:
        trip_id = int(trip_id_str)
        try:
            trip = Trip.objects.get(pk=trip_id)
            # Знаходимо WorkRequestItems, пов'язані з цим відрядженням та обраними ОІДами
            # Це допоможе визначити тип робіт (Атестація/ІК) для кожного ОІД у контексті відрядження
            
            # Отримуємо всі документи для обраних ОІДів
            relevant_documents = Document.objects.filter(
                oid_id__in=oid_ids
            ).select_related('document_type', 'oid__unit').order_by('oid__cipher', '-work_date')

            # Фільтруємо документи за логікою:
            # - Якщо ІК: всі документи до "Висновок ІК" включно.
            # - Якщо Атестація: пакет з "Акт атестації", тільки якщо він зареєстрований.
            
            # Для визначення типу робіт, нам потрібен зв'язок Trip -> WorkRequest -> WorkRequestItem -> OID
            # Це може бути складним запитом. Спрощений підхід:
            # Ми знаємо ОІДи. Для кожного ОІД шукаємо WorkRequestItem, який пов'язаний з цим відрядженням.
            
            # Збираємо work_types для кожного обраного ОІД у цьому відрядженні
            oid_work_types = {} # {oid_id: work_type}
            work_request_items_in_trip = WorkRequestItem.objects.filter(
                request__trips=trip, # Зв'язок через M2M Trip <-> WorkRequest
                oid_id__in=oid_ids
            ).select_related('oid', 'request')

            for wri in work_request_items_in_trip:
                if wri.oid_id not in oid_work_types: # Беремо перший знайдений тип роботи для ОІД в рамках відрядження
                    oid_work_types[wri.oid_id] = wri.work_type

            # Визначаємо DocumentType для "Акт атестації" та "Висновок ІК"
            # Краще мати константи або slug для цих типів
            attestation_act_type = DocumentType.objects.filter(name__icontains="Акт атестації").first()
            ik_conclusion_type = DocumentType.objects.filter(duration_months="20").first()
            
            # Інші типи документів, які входять до пакетів (за потреби)
            # attestation_act_type = DocumentType.objects.filter(name__icontains="Акт атестації").first()
            # ik_conclusion_type = DocumentType.objects.filter(name__icontains="Висновок ІК").first()
            # program_method_type = DocumentType.objects.filter(name__icontains="Програма і методика").first()
            # plan_search_type = DocumentType.objects.filter(name__icontains="План пошуку ЗП").first()
            # act_search_type = DocumentType.objects.filter(name__icontains="Акт пошуку ЗП").first()
            # protocol1a_type = ... protocol2a_type = ... protocol_ik_type = ...
            # prescription_type = DocumentType.objects.filter(name__icontains="Припис").first()


            for doc in relevant_documents:
                work_type_for_this_oid = oid_work_types.get(doc.oid_id)
                
                add_document = False
                if work_type_for_this_oid == WorkTypeChoices.IK:
                    # Для ІК додаємо всі документи, що традиційно входять до пакету
                    # Або, якщо простіше, тільки Висновок ІК (або всі до нього)
                    # Наприклад, якщо doc.document_type є одним з документів пакету ІК
                    if doc.document_type == ik_conclusion_type: # Або більш розширена логіка
                        add_document = True
                    # Тут можна додати інші типи документів для ІК: План пошуку, Акт пошуку, Протоколи...
                    # if doc.document_type in [plan_search_type, act_search_type, protocol_ik_type, ik_conclusion_type]:
                    # add_document = True

                elif work_type_for_this_oid == WorkTypeChoices.ATTESTATION:
                    # Для Атестації додаємо пакет документів, якщо Акт Атестації зареєстрований
                    if doc.document_type == attestation_act_type and \
                       doc.dsszzi_registered_number and doc.dsszzi_registered_date:
                        add_document = True # Додаємо сам зареєстрований акт
                        # Тут можна додати логіку для додавання інших документів з пакету атестації,
                        # якщо головний акт зареєстрований.
                        # Наприклад:
                        # if doc.document_type in [program_method_type, plan_search_type, ... , attestation_act_type, prescription_type]:
                        #    # Потрібно перевірити, чи основний акт для цього ОІД/роботи зареєстрований
                        #    main_att_act = Document.objects.filter(oid=doc.oid, document_type=attestation_act_type, work_request_item__work_type=WorkTypeChoices.ATTESTATION, dsszzi_registered_number__isnull=False).exists()
                        #    if main_att_act:
                        #        add_document = True

                if add_document:
                    documents_data.append({
                        'id': doc.id,
                        'text': f"в/ч {doc.oid.unit.code} ОІД: {doc.oid.cipher} | Документ: {doc.document_type.name} підг.№ {doc.document_number} від {doc.work_date.strftime('%d.%m.%Y')}"
                    })
        except Trip.DoesNotExist:
            pass
        except Exception as e:
            print(f"Error in ajax_load_documents_for_trip_oids: {e}")
            return JsonResponse({'error': str(e)}, status=500)
            
    return JsonResponse(documents_data, safe=False)

def get_last_document_expiration_date(oid_instance, document_name_keyword, work_type_choice=None):
    try:
        doc_type_filters = Q(name__icontains=document_name_keyword)
        if work_type_choice:
            doc_type_filters &= Q(work_type=work_type_choice)
        
        relevant_doc_type_qs = DocumentType.objects.filter(doc_type_filters)
        if not relevant_doc_type_qs.exists():
            return None
        relevant_doc_type = relevant_doc_type_qs.first()

        last_document = Document.objects.filter(
            oid=oid_instance, # Використовуємо переданий екземпляр OID
            document_type=relevant_doc_type,
            expiration_date__isnull=False
        ).order_by('-work_date', '-doc_process_date').first()
        
        return last_document.expiration_date if last_document else None
    except Exception as e:
        print(f"Помилка get_last_document_expiration_date для ОІД {oid_instance.cipher if oid_instance else 'N/A'} ({document_name_keyword}): {e}")
        return None

	#  логіка оновлення статусу заявки:
def check_and_update_status_based_on_documents(self):
    """
    Перевіряє, чи виконані умови для зміни статусу цього WorkRequestItem,
    ґрунтуючись на наявності та стані документів.
    
    Логіка статусів:
    - TO_SEND_AA: Документи опрацьовано для Атестації, готові до відправки в ДССЗЗІ
    - ON_REGISTRATION: Документи відправлено на реєстрацію в ДССЗЗІ (для Атестації)
    - TO_SEND_VCH: Документи зареєстровано в ДССЗЗІ / опрацьовано для ІК, готові до відправки у в/ч
    - COMPLETED: Документи відправлено у в/ч
    """
    print(f"[WRI_STATUS_CHECKER] Checking status for WRI ID {self.id} (OID: {self.oid.cipher}, Work Type: {self.work_type})")

    if self.status in [WorkRequestStatusChoices.COMPLETED, WorkRequestStatusChoices.CANCELED]:
        print(f"[WRI_STATUS_CHECKER] WRI ID {self.id} is already {self.status}. No update needed.")
        return

    existing_docs_for_item = Document.objects.filter(work_request_item=self)
    
    fields_to_update = []
    new_status = self.status
    document_date = None

    # --- ОНОВЛЕНА ЛОГІКА ДЛЯ АТЕСТАЦІЇ (ATTESTATION та PLAND_ATTESTATION) ---
    if self.work_type in [WorkTypeChoices.PLAND_ATTESTATION, WorkTypeChoices.ATTESTATION]:
        attestation_act_type = DocumentType.objects.filter(name__icontains="Акт атестації").first()
        
        if not attestation_act_type:
            print(f"[WRI_STATUS_CHECKER] Attestation Act type not found in system!")
            return
        
        # Шукаємо Акт Атестації для цього WRI
        attestation_doc = existing_docs_for_item.filter(
            document_type=attestation_act_type
        ).order_by('-doc_process_date').first()
        
        if attestation_doc:
            # Перевіряємо, чи є реєстраційний номер ДССЗЗІ
            has_registration = (
                attestation_doc.dsszzi_registered_number and 
                attestation_doc.dsszzi_registered_number.strip() != ''
            )
            
            # Перевіряємо, чи документ відправлено на реєстрацію (має attestation_registration_sent)
            is_sent_for_registration = attestation_doc.attestation_registration_sent is not None
            
            # НОВИНКА: Перевіряємо, чи документ відправлено у в/ч (має trip_result_sent)
            is_sent_to_unit = hasattr(attestation_doc, 'trip_result_sent') and attestation_doc.trip_result_sent is not None
            
            if is_sent_to_unit:
                # Документ відправлено у в/ч → статус "Виконано"
                new_status = WorkRequestStatusChoices.COMPLETED
                document_date = attestation_doc.doc_process_date
                print(f"[WRI_STATUS_CHECKER] Attestation doc sent to unit. Setting status to COMPLETED.")
                
            elif has_registration:
                # Є реєстраційний номер, але ще не відправлено у в/ч → "Готово до відправки в в/ч"
                new_status = WorkRequestStatusChoices.TO_SEND_VCH
                document_date = attestation_doc.dsszzi_registered_date or attestation_doc.doc_process_date
                print(f"[WRI_STATUS_CHECKER] Attestation doc has registration number but not sent to unit. Setting status to TO_SEND_VCH.")
                
            elif is_sent_for_registration:
                # Відправлено на реєстрацію, але номера ще немає → "На реєстрації в ДССЗЗІ"
                new_status = WorkRequestStatusChoices.ON_REGISTRATION
                document_date = attestation_doc.doc_process_date
                print(f"[WRI_STATUS_CHECKER] Attestation doc sent for registration. Setting status to ON_REGISTRATION.")
                
            elif attestation_doc.doc_process_date:
                # Документ опрацьовано, але не відправлено → "Готово до відправки в ДССЗЗІ"
                new_status = WorkRequestStatusChoices.TO_SEND_AA
                document_date = attestation_doc.doc_process_date
                print(f"[WRI_STATUS_CHECKER] Attestation doc processed but not sent. Setting status to TO_SEND_AA.")
        else:
            print(f"[WRI_STATUS_CHECKER] No Attestation Act found for WRI ID {self.id}.")

    # --- ЛОГІКА ДЛЯ ІК (БЕЗ ЗМІН) ---
    elif self.work_type == WorkTypeChoices.IK:
        ik_conclusion_type = DocumentType.objects.filter(duration_months=20).first()
        
        if not ik_conclusion_type:
            print(f"[WRI_STATUS_CHECKER] IK Conclusion type not found in system!")
            return
        
        # Шукаємо Висновок ІК для цього WRI
        ik_doc = existing_docs_for_item.filter(
            document_type=ik_conclusion_type
        ).order_by('-doc_process_date').first()
        
        if ik_doc:
            # Перевіряємо, чи документ відправлено у в/ч (має trip_result_sent)
            is_sent_to_unit = hasattr(ik_doc, 'trip_result_sent') and ik_doc.trip_result_sent is not None
            
            if is_sent_to_unit:
                # Документ відправлено у в/ч → "Виконано"
                new_status = WorkRequestStatusChoices.COMPLETED
                document_date = ik_doc.doc_process_date
                print(f"[WRI_STATUS_CHECKER] IK doc sent to unit. Setting status to COMPLETED.")
                
            elif ik_doc.doc_process_date:
                # Документ опрацьовано, але не відправлено → "Готово до відправки в в/ч"
                new_status = WorkRequestStatusChoices.TO_SEND_VCH
                document_date = ik_doc.doc_process_date
                print(f"[WRI_STATUS_CHECKER] IK doc processed but not sent. Setting status to TO_SEND_VCH.")
        else:
            print(f"[WRI_STATUS_CHECKER] No IK Conclusion found for WRI ID {self.id}.")

    # --- Оновлюємо статус, якщо він змінився ---
    if new_status != self.status:
        self.status = new_status
        fields_to_update.append('status')
        print(f"[WRI_STATUS_CHECKER] Status changed from {self.status} to {new_status}")
    
    # --- Оновлюємо дату фактичного опрацювання ---
    if document_date and not self.docs_actually_processed_on:
        self.docs_actually_processed_on = document_date
        fields_to_update.append('docs_actually_processed_on')
        print(f"[WRI_STATUS_CHECKER] Setting docs_actually_processed_on to {document_date}")
    
    # --- Зберігаємо зміни ---
    if fields_to_update:
        fields_to_update.append('updated_at')
        self.save(update_fields=fields_to_update)
        print(f"[WRI_STATUS_CHECKER] WRI ID {self.id} updated. New status: {self.get_status_display()}")
        
        # Оновлюємо статус батьківської заявки
        self.update_parent_request_status()
    else:
        print(f"[WRI_STATUS_CHECKER] No changes needed for WRI ID {self.id}. Current status: {self.get_status_display()}")


        # ... (решта методів моделі WorkRequestItem, наприклад update_parent_request_status) ...
    
def update_parent_request_status(self):
	"""
	Оновлює статус батьківської заявки WorkRequest на основі статусів
	всіх її елементів WorkRequestItem.
	
	Логіка:
	- Якщо всі COMPLETED → заявка COMPLETED
	- Якщо всі CANCELED → заявка CANCELED
	- Якщо є хоча б один ON_REGISTRATION → заявка ON_REGISTRATION
	- Якщо є хоча б один TO_SEND_AA або TO_SEND_VCH → відповідний статус заявки
	- Якщо є хоча б один IN_PROGRESS → заявка IN_PROGRESS
	- Інакше → PENDING
	"""
	work_request = self.request
	all_items = work_request.items.all()

	if not all_items.exists():
		print(f"[REQUEST_STATUS_UPDATER] WorkRequest ID {work_request.id} has no items.")
		if work_request.status != WorkRequestStatusChoices.PENDING:
			work_request.status = WorkRequestStatusChoices.PENDING
			work_request.save(update_fields=['status', 'updated_at'])
			print(f"[REQUEST_STATUS_UPDATER] Set to PENDING (no items).")
		return

	print(f"[REQUEST_STATUS_UPDATER] Updating status for WorkRequest ID {work_request.id}")
	
	# Підраховуємо статуси елементів
	status_counts = {
		'completed': all_items.filter(status=WorkRequestStatusChoices.COMPLETED).count(),
		'canceled': all_items.filter(status=WorkRequestStatusChoices.CANCELED).count(),
		'on_registration': all_items.filter(status=WorkRequestStatusChoices.ON_REGISTRATION).count(),
		'to_send_aa': all_items.filter(status=WorkRequestStatusChoices.TO_SEND_AA).count(),
		'to_send_vch': all_items.filter(status=WorkRequestStatusChoices.TO_SEND_VCH).count(),
		'in_progress': all_items.filter(status=WorkRequestStatusChoices.IN_PROGRESS).count(),
		'pending': all_items.filter(status=WorkRequestStatusChoices.PENDING).count(),
		'total': all_items.count()
	}
	
	print(f"[REQUEST_STATUS_UPDATER] Status counts: {status_counts}")
	
	original_status = work_request.status
	new_status = original_status

	# Визначаємо новий статус заявки за пріоритетом
	if status_counts['completed'] == status_counts['total']:
		# Всі елементи виконані
		new_status = WorkRequestStatusChoices.COMPLETED
		print(f"[REQUEST_STATUS_UPDATER] All items COMPLETED.")
		
	elif status_counts['canceled'] == status_counts['total']:
		# Всі елементи скасовані
		new_status = WorkRequestStatusChoices.CANCELED
		print(f"[REQUEST_STATUS_UPDATER] All items CANCELED.")
		
	elif status_counts['on_registration'] > 0:
		# Є елементи на реєстрації в ДССЗЗІ
		new_status = WorkRequestStatusChoices.ON_REGISTRATION
		print(f"[REQUEST_STATUS_UPDATER] Has items ON_REGISTRATION.")
		
	elif status_counts['to_send_aa'] > 0 and status_counts['to_send_vch'] > 0:
		# Є елементи обох типів, готові до відправки - встановлюємо статус за першим знайденим
		# Або можна створити окремий статус "Готово до відправки"
		new_status = WorkRequestStatusChoices.TO_SEND_AA
		print(f"[REQUEST_STATUS_UPDATER] Has items TO_SEND (both types).")
		
	elif status_counts['to_send_aa'] > 0:
		# Є елементи, готові до відправки в ДССЗЗІ
		new_status = WorkRequestStatusChoices.TO_SEND_AA
		print(f"[REQUEST_STATUS_UPDATER] Has items TO_SEND_AA.")
		
	elif status_counts['to_send_vch'] > 0:
		# Є елементи, готові до відправки у в/ч
		new_status = WorkRequestStatusChoices.TO_SEND_VCH
		print(f"[REQUEST_STATUS_UPDATER] Has items TO_SEND_VCH.")
		
	elif status_counts['in_progress'] > 0:
		# Є елементи в роботі
		new_status = WorkRequestStatusChoices.IN_PROGRESS
		print(f"[REQUEST_STATUS_UPDATER] Has items IN_PROGRESS.")
		
	elif status_counts['pending'] > 0:
		# Є елементи, що очікують
		new_status = WorkRequestStatusChoices.PENDING
		print(f"[REQUEST_STATUS_UPDATER] Has items PENDING.")

	# Зберігаємо новий статус, якщо він змінився
	if original_status != new_status:
		work_request.status = new_status
		work_request.save(update_fields=['status', 'updated_at'])
		print(f"[REQUEST_STATUS_UPDATER] WorkRequest ID {work_request.id} status changed: {original_status} → {new_status}")
	else:
		print(f"[REQUEST_STATUS_UPDATER] WorkRequest ID {work_request.id} status unchanged: {new_status}")


def update_request_status(self):
	"""
	Оновлює статус батьківської заявки WorkRequest на основі статусів
	всіх її елементів WorkRequestItem.
	"""
	work_request = self.request
	all_items = work_request.items.all()

	if not all_items.exists():
		# Якщо заявка не має елементів (наприклад, щойно створена і ще не додані, або всі видалені)
		# Можна встановити PENDING або залишити як є, залежно від бізнес-логіки.
		# Якщо це відбувається після видалення останнього елемента, можливо, заявку треба скасувати або повернути в PENDING.
		if work_request.status != WorkRequestStatusChoices.PENDING: # Якщо вже не PENDING
			work_request.status = WorkRequestStatusChoices.PENDING # або інший логічний статус
			work_request.save(update_fields=['status', 'updated_at'])
		return

	# Перевіряємо, чи всі елементи завершені або скасовані
	is_all_items_processed = all(
		item.status in [WorkRequestStatusChoices.COMPLETED, WorkRequestStatusChoices.CANCELED]
		for item in all_items
	)

	original_request_status = work_request.status
	new_request_status = original_request_status # За замовчуванням не змінюємо


	if is_all_items_processed:
		# Якщо всі елементи оброблені, визначаємо фінальний статус заявки
		if all_items.filter(status=WorkRequestStatusChoices.COMPLETED).exists():
			# Якщо є хоча б один виконаний елемент, заявка вважається виконаною
			new_request_status = WorkRequestStatusChoices.COMPLETED
		elif all_items.filter(status=WorkRequestStatusChoices.CANCELED).count() == all_items.count():
			# Якщо всі елементи скасовані (і немає виконаних)
			new_request_status = WorkRequestStatusChoices.CANCELED
		else:
			# Ситуація, коли всі CANCELED, але був хоча б один COMPLETED, вже покрита першою умовою.
			# Якщо всі CANCELED і не було COMPLETED - це друга умова.
			# Якщо є суміш COMPLETED і CANCELED, то COMPLETED має пріоритет.
			# Якщо логіка інша (напр. "Частково виконано"), її треба додати.
				new_request_status = WorkRequestStatusChoices.COMPLETED # За замовчуванням для змішаних processed
	else:
		# Якщо не всі елементи оброблені, перевіряємо наявність "В роботі" або "Очікує"
		if all_items.filter(status=WorkRequestStatusChoices.IN_PROGRESS).exists():
			new_request_status = WorkRequestStatusChoices.IN_PROGRESS
		elif all_items.filter(status=WorkRequestStatusChoices.PENDING).exists():
			new_request_status = WorkRequestStatusChoices.PENDING
		# Можливий випадок: немає IN_PROGRESS, немає PENDING, але не всі processed.
		# Це може статися, якщо є власні статуси. Для стандартних це малоймовірно.
		# У такому разі, можна залишити поточний статус заявки або встановити PENDING.
	if original_request_status != new_request_status:
		work_request.status = new_request_status
		work_request.save(update_fields=['status', 'updated_at'])
		print(f"[DEBUG] WorkRequest ID {work_request.id} status successfully saved as '{work_request.get_status_display()}'.")
	else:
		print(f"[DEBUG] WorkRequest ID {work_request.id} status '{work_request.get_status_display()}' remains unchanged.")
	print(f"--- [DEBUG] WRI.update_request_status() FINISHED for WRI ID: {self.id} ---")

def __str__(self):
	# return f"{self.oid.cipher} - {self.get_work_type_display()} ({self.get_status_display()})"
	return f"ОІД: {self.oid.cipher} ({self.oid.oid_type}) - Робота: {self.get_work_type_display()} (Статус: {self.status})"

class Meta:
	unique_together = ('request', 'oid', 'work_type') # Один ОІД не може мати двічі одну і ту ж роботу в одній заявці
	verbose_name = "Заявки: Елемент заявки (ОІД)"
	verbose_name_plural = "Заявки: Елементи заявки (ОІД)"
	# ordering = ['request', 'oid'] # Додано сортування
	ordering = ['request', 'request__incoming_date' ] # Додав сортування за замовчуванням


def main_dashboard(request):
    """
    Головна сторінка. Фільтрація ВЧ -> ОІД відбувається на сервері 
    через перезавантаження сторінки.
    """
    try:
        add_request_url = reverse('oids:add_work_request') 
        plan_trip_url = reverse('oids:plan_trip_view_name')
        add_document_processing_url = reverse('oids:add_document_processing_view_name')
    except Exception:
        add_request_url, plan_trip_url, add_document_processing_url = '#', '#', '#'

    all_units = Unit.objects.select_related('territorial_management').order_by('name')
    
    selected_unit_id_str = request.GET.get('unit')
    selected_unit_object = None
    
    oids_creating_list = []
    oids_active_list = []
    oids_cancelled_list = []

    if selected_unit_id_str:
        try:
            selected_unit_id = int(selected_unit_id_str)
            selected_unit_object = Unit.objects.get(pk=selected_unit_id)
            
            oids_for_selected_unit = OID.objects.filter(unit_id=selected_unit_id)\
                                              .select_related('unit')\
                                              .order_by('cipher')

            for oid_instance in oids_for_selected_unit:
                oid_item_data = {
                    'id': oid_instance.id,
                    'cipher': oid_instance.cipher,
                    'full_name': oid_instance.full_name or oid_instance.unit.city,
                    'oid_type_display': oid_instance.get_oid_type_display(),
                    'status_display': oid_instance.get_status_display(),
                    'detail_url': reverse('oids:oid_detail_view_name', args=[oid_instance.id])
                }
                OID_to_show_main_dashboard_creating = [
                    OIDStatusChoices.NEW,
                    OIDStatusChoices.RECEIVED_TZ,
                    OIDStatusChoices.RECEIVED_TZ_REPEAT,
                    OIDStatusChoices.RECEIVED_TZ_APPROVE,
                    OIDStatusChoices.RECEIVED_REQUEST,
                    OIDStatusChoices.RECEIVED_REQUEST_ATTESTATION,
                    OIDStatusChoices.ATTESTED,
                    OIDStatusChoices.AZR_SEND , 
                    OIDStatusChoices.RECEIVED_DECLARATION
                    ]
                OID_to_show_main_dashboard_active = [
                    OIDStatusChoices.ACTIVE,
                    OIDStatusChoices.RECEIVED_REQUEST_IK,
                    OIDStatusChoices.RECEIVED_REQUEST_PLAND_ATTESTATION,
				]
                OID_to_show_main_dashboard_cancel = [
                    OIDStatusChoices.CANCELED,
                    OIDStatusChoices.TERMINATED,
                    OIDStatusChoices.INACTIVE
				]
                if oid_instance.status in OID_to_show_main_dashboard_creating:
                    oids_creating_list.append(oid_item_data)
                elif oid_instance.status in OID_to_show_main_dashboard_active:
                    oid_item_data['ik_expiration_date'] = get_last_document_expiration_date(oid_instance, 'Висновок')
                    oid_item_data['attestation_expiration_date'] = get_last_document_expiration_date(oid_instance, 'Акт атестації')
                    oid_item_data['prescription_expiration_date'] = get_last_document_expiration_date(oid_instance, 'Припис')
                    oids_active_list.append(oid_item_data)
                elif oid_instance.status in OID_to_show_main_dashboard_cancel:
                    oids_cancelled_list.append(oid_item_data)
        except (ValueError, Unit.DoesNotExist):
            selected_unit_object = None 
            # Можна додати повідомлення про помилку, якщо unit_id невірний
            # messages.error(request, "Обрана військова частина не знайдена.")

    context = {
        'add_request_url': add_request_url,
        'plan_trip_url': plan_trip_url,
        'add_document_processing_url': add_document_processing_url,
        'all_units': all_units,
        'selected_unit_id': selected_unit_id_str, # Передаємо рядок для порівняння в шаблоні
        'selected_unit_object': selected_unit_object,
        'oids_creating': oids_creating_list,
        'oids_active': oids_active_list,
        'oids_cancelled': oids_cancelled_list,
    }
    return render(request, 'oids/main_dashboard.html', context)

@login_required 
@transaction.atomic
def send_trip_results_view(request):
    if request.method == 'POST':
        form = TripResultSendForm(request.POST)
        if form.is_valid():
            # Метод save() форми тепер обробляє збереження M2M та оновлення статусів
            trip_result = form.save() 
            messages.success(request, f"Результати відрядження (для відправки до ВЧ) успішно збережено.")
            # Перенаправлення на список результатів або на деталі цього запису
            return redirect('oids:list_trip_results_for_units') # Потрібно створити цей URL та view
        else:
            messages.error(request, "Будь ласка, виправте помилки у формі.")
            print(f"TripResultSendForm errors: {form.errors.as_json()}")
    else:
        form = TripResultSendForm() # Порожня форма для GET-запиту

    context = {
        'form': form,
        'page_title': 'Сформувати пакет документів (результати відрядження) для відправки до ВЧ'
    }
    return render(request, 'oids/forms/send_trip_results_form.html', context)


# ... (решта ваших views: oid_detail_view, форми для додавання тощо) ...
@login_required # Не забудьте додати oid_detail_view з попередньої відповіді.
def oid_detail_view(request, oid_id):
    oid = get_object_or_404(
        OID.objects.select_related(
            'unit',
            'unit__territorial_management'
        ),
        pk=oid_id
    )
    

		# Отримуємо інформацію про активний процес для цього ОІД
    active_process = oid.active_process if hasattr(oid, 'active_process') else None
    step_instances = []
    next_actionable_step = None

    if active_process:
        step_instances = active_process.step_instances.all().order_by('process_step__order')
        # Знаходимо перший крок, який ще не виконано
        next_actionable_step = step_instances.filter(status='очікує').first()
        

    status_changes = oid.status_changes.select_related(
        'initiating_document__document_type', # initiating_document може бути null
        'changed_by'
    ).order_by('-changed_at')

    work_requests_for_oid = WorkRequest.objects.filter(
        items__oid=oid
    ).select_related('unit').prefetch_related(
        Prefetch('items', queryset=WorkRequestItem.objects.filter(oid=oid).select_related('oid'))
    ).distinct().order_by('-incoming_date')
    
    technical_tasks = oid.technical_tasks.select_related('reviewed_by').order_by('-input_date')

    documents = oid.documents.select_related(
        'document_type',
        'author',
        'work_request_item__request',
        'attestation_registration_sent' # Додаємо для доступу до даних відправки
    ).order_by('-doc_process_date', '-work_date')

    # Отримуємо тільки документи типу "Акт атестації" для цього ОІД для секції реєстрації
    attestation_acts_for_oid = documents.filter(document_type__name__icontains="Акт атестації")

    trips_for_oid = oid.trips.prefetch_related(
        'units',
        'persons',
        'work_requests'
    ).order_by('-start_date')
    
    # Видаляємо логіку, пов'язану з AttestationItem, оскільки її більше немає
    # attestation_registrations_for_oid = AttestationRegistration.objects.filter(
    #     items_sent__attestation_document__oid=oid # Оновлений шлях доступу через Document
    # ).prefetch_related(
    #     Prefetch('items_sent', queryset=Document.objects.filter(oid=oid, document_type__duration_months=60).select_related('document_type')),
    #     'response_received' # related_name з AttestationResponse
    # ).distinct().order_by('-outgoing_letter_date')
    # Натомість, ми будемо використовувати attestation_acts_for_oid та їх зв'язок з AttestationRegistration


    last_attestation_expiration = get_last_document_expiration_date(oid, 'Акт атестації')
    last_ik_expiration = get_last_document_expiration_date(oid, 'Висновок')
    last_prescription_expiration = get_last_document_expiration_date(oid, 'Припис')

    context = {
        'oid': oid,
        'status_changes': status_changes,
        'work_requests_for_oid': work_requests_for_oid,
        # 'work_request_items_for_oid': work_request_items, # Це тепер всередині work_requests_for_oid
        'technical_tasks': technical_tasks,
        'documents': documents, # Всі документи
        'attestation_acts_for_oid': attestation_acts_for_oid, # Тільки акти атестації для цього ОІД
        'trips_for_oid': trips_for_oid,
        # 'attestation_registrations': attestation_registrations_for_oid, # Видалено, використовуємо attestation_acts_for_oid
        'last_attestation_expiration': last_attestation_expiration,
        'last_ik_expiration': last_ik_expiration,
        'last_prescription_expiration': last_prescription_expiration,
        'active_process': active_process,
        'step_instances': step_instances,
        'next_actionable_step': next_actionable_step,
    }
    return render(request, 'oids/oid_detail.html', context)


# ... (main_dashboard, ajax_load_oids_for_unit_categorized, ajax_load_oids_for_unit, oid_detail_view) ...
# Переконайся, що функція get_last_document_expiration_date визначена вище


@login_required
def oid_create_view(request):
    if request.method == 'POST':
        form = OIDCreateForm(request.POST)
        if form.is_valid():
            oid_instance = form.save()
            messages.success(request, f'ОІД "{oid_instance.cipher}" успішно створено.')
            return redirect('oids:oid_detail_view_name', oid_id=oid_instance.id) # Перенаправляємо на деталі нового ОІД
        else:
            messages.error(request, 'Будь ласка, виправте помилки у формі.')
    else: # GET request
        form = OIDCreateForm()
    
    context = {
        'form': form,
        'page_title': 'Створення нового Об\'єкта Інформаційної Діяльності'
    }
    return render(request, 'oids/forms/oid_create_form.html', context)


@login_required 
@transaction.atomic
def add_work_request_view(request):
    """
    Створення нової заявки на проведення робіт
    """
    # Отримуємо екземпляр Unit, якщо ID передано в GET-запиті
    unit_instance = None
    unit_id_from_get = request.GET.get('unit')
    if unit_id_from_get:
        try:
            unit_instance = Unit.objects.get(pk=unit_id_from_get)
        except Unit.DoesNotExist:
            messages.error(request, "❌ Обрану військову частину не знайдено.")

    initial_main_form = {}
    if unit_instance:
        initial_main_form['unit'] = unit_instance

    if request.method == 'POST':
        main_form = WorkRequestForm(request.POST, request.FILES, prefix='main')
        
        # Передаємо екземпляр unit у form_kwargs для кожної форми у формсеті
        selected_unit_id = request.POST.get('main-unit')
        parent_unit_for_formset = None
        if selected_unit_id:
            try:
                parent_unit_for_formset = Unit.objects.get(pk=selected_unit_id)
            except Unit.DoesNotExist:
                pass
        
        formset = WorkRequestItemFormSet(
            request.POST, 
            request.FILES, 
            prefix='items',
            form_kwargs={'parent_instance_unit': parent_unit_for_formset}
        )

        if main_form.is_valid() and formset.is_valid():
            try:
                # Зберігаємо основну заявку
                work_request_instance = main_form.save()
                
                # Прив'язуємо формсет до збереженої заявки
                formset.instance = work_request_instance
                saved_items = formset.save()
                
                # Збираємо статистику для повідомлення
                work_types_count = {}
                oids_list = []
                
                for item in saved_items:
                    # Рахуємо типи робіт
                    work_type_display = item.get_work_type_display()
                    if work_type_display not in work_types_count:
                        work_types_count[work_type_display] = 0
                    work_types_count[work_type_display] += 1
                    
                    # Збираємо інфо про ОІДи
                    oids_list.append({
                        'cipher': item.oid.cipher,
                        'unit_code': item.oid.unit.code if item.oid.unit else 'Без ВЧ',
                        'work_type': work_type_display
                    })
                
                # Формуємо детальне повідомлення
                success_message = (
                    f"✅ <strong>Заявку успішно створено!</strong><br>"
                    f"📋 Номер заявки: <strong>№{work_request_instance.incoming_number}</strong> "
                    f"від <strong>{work_request_instance.incoming_date.strftime('%d.%m.%Y')}</strong><br>"
                )
                
                # Додаємо інформацію про військову частину
                if work_request_instance.unit:
                    success_message += f"🏢 Військова частина: <strong>{work_request_instance.unit.code}</strong><br>"
                
                success_message += (
                    f"📝 ОІДів у заявці: <strong>{len(saved_items)}</strong><br>"
                    f"<br><strong>Типи робіт:</strong><br>"
                )
                
                # Виводимо статистику по типах робіт
                for work_type, count in work_types_count.items():
                    success_message += f"&nbsp;&nbsp;&nbsp;• {work_type}: <strong>{count}</strong><br>"
                
                success_message += "<br><strong>Деталі по ОІДам:</strong><br>"
                
                # Виводимо кожен ОІД
                for oid_info in oids_list:
                    success_message += (
                        f"&nbsp;&nbsp;&nbsp;• ВЧ <strong>{oid_info['unit_code']}</strong> | "
                        f"ОІД <strong>{oid_info['cipher']}</strong> | "
                        f"Робота: {oid_info['work_type']}<br>"
                    )
                
                messages.success(request, mark_safe(success_message))
                
                # Додаткове інформаційне повідомлення
                messages.info(
                    request,
                    mark_safe(
                        f"ℹ️ Наступний крок: <a href=\"{reverse('oids:work_request_detail', args=[work_request_instance.id])}\">переглянути деталі заявки</a> "
                        f"або <a href=\"{reverse('oids:list_work_requests')}\">повернутись до списку заявок</a>."
                    )
                )
                
                return redirect('oids:list_work_requests')
                
            except Exception as e:
                # Обробка помилок
                messages.error(
                    request,
                    mark_safe(
                        f"❌ <strong>Помилка при створенні заявки:</strong><br>"
                        f"{str(e)}<br><br>"
                        f"Зміни не були збережені. Спробуйте ще раз."
                    )
                )
                return redirect('oids:add_work_request')
        
        else:
            # Якщо форми не валідні - показуємо детальні помилки
            error_messages = []
            
            # Помилки основної форми
            if not main_form.is_valid():
                error_messages.append("<strong>Помилки в загальній інформації:</strong>")
                
                if main_form.non_field_errors():
                    for error in main_form.non_field_errors():
                        error_messages.append(f"• {error}")
                
                for field_name, errors in main_form.errors.items():
                    if field_name == '__all__':
                        continue
                    field_label = main_form.fields[field_name].label or field_name
                    for error in errors:
                        error_messages.append(f"• {field_label}: {error}")
            
            # Помилки формсету
            if not formset.is_valid():
                error_messages.append("<br><strong>Помилки в елементах заявки (ОІДи та роботи):</strong>")
                
                # Загальні помилки формсету
                if formset.non_form_errors():
                    for error in formset.non_form_errors():
                        error_messages.append(f"• {error}")
                
                # Помилки окремих форм
                for i, form_errors in enumerate(formset.errors):
                    if form_errors:
                        error_messages.append(f"<br><strong>ОІД/Робота #{i+1}:</strong>")
                        for field, errors in form_errors.items():
                            if field == '__all__':
                                for error in errors:
                                    error_messages.append(f"• {error}")
                            else:
                                field_label = formset.forms[i].fields.get(field).label if field in formset.forms[i].fields else field
                                for error in errors:
                                    error_messages.append(f"• {field_label}: {error}")
            
            messages.error(
                request,
                mark_safe(
                    "❌ <strong>Будь ласка, виправте помилки у формі:</strong><br>" + 
                    "<br>".join(error_messages)
                )
            )
    
    else:  # GET request
        main_form = WorkRequestForm(initial=initial_main_form, prefix='main')
        formset = WorkRequestItemFormSet(prefix='items', form_kwargs={'parent_instance_unit': unit_instance})

    context = {
        'main_form': main_form,
        'formset': formset,
        'page_title': 'Створення нової заявки на проведення робіт',
        'selected_unit_instance': unit_instance,
    }
    return render(request, 'oids/forms/add_work_request_form.html', context)


# ДОДАТКОВА ФУНКЦІЯ: Статистика про заявку
def get_work_request_stats(work_request):
    """
    Повертає детальну статистику про заявку
    Можна використовувати для додаткових звітів
    """
    items = work_request.items.all()
    
    stats = {
        'total_items': items.count(),
        'by_work_type': {},
        'by_status': {},
        'oids_list': []
    }
    
    for item in items:
        # Статистика по типу робіт
        work_type = item.get_work_type_display()
        if work_type not in stats['by_work_type']:
            stats['by_work_type'][work_type] = 0
        stats['by_work_type'][work_type] += 1
        
        # Статистика по статусу
        status = item.get_status_display()
        if status not in stats['by_status']:
            stats['by_status'][status] = 0
        stats['by_status'][status] += 1
        
        # Список ОІДів
        stats['oids_list'].append({
            'cipher': item.oid.cipher,
            'work_type': work_type,
            'status': status,
            'unit_code': item.oid.unit.code if item.oid.unit else None
        })
    
    return stats

@login_required 
def work_request_detail_view(request, pk):
    """
    Представлення для відображення деталей заявки на роботи
    та оновлення статусу її окремих пунктів.
    """
    # Отримуємо об'єкт заявки або повертаємо помилку 404, якщо його не існує
    work_request = get_object_or_404(WorkRequest, pk=pk)

    if request.method == 'POST':
        # Отримуємо ID пункту та новий статус із POST-запиту
        item_id = request.POST.get('item_id')
        new_status = request.POST.get('status')

        if item_id and new_status:
            try:
                # Знаходимо конкретний пункт, що належить цій заявці
                item_to_update = WorkRequestItem.objects.get(pk=item_id, request=work_request)
                
                # Перевіряємо, чи є такий статус у моделі
                valid_statuses = [status[0] for status in WorkRequestStatusChoices.choices]
                if new_status in valid_statuses:
                    item_to_update.status = new_status
                    item_to_update.save()
                    messages.success(request, f'Статус для ОІД "{item_to_update.oid}" оновлено на "{item_to_update.get_status_display()}".')
                else:
                    messages.error(request, 'Невірний статус.')

            except WorkRequestItem.DoesNotExist:
                messages.error(request, 'Пункт заявки не знайдено.')
        else:
            messages.warning(request, 'Необхідно надати ID пункту та статус.')
        
        # Перенаправляємо користувача на цю ж сторінку, щоб уникнути повторної відправки форми
        return redirect('oids:work_request_detail', pk=work_request.pk)

    # Для GET-запиту просто готуємо дані для шаблону
    # Отримуємо всі пункти, пов'язані з цією заявкою
    items = work_request.items.all().select_related('oid')
    
    # Отримуємо доступні статуси з моделі для випадаючого списку в шаблоні
    status_choices = WorkRequestStatusChoices.choices

    context = {
        'work_request': work_request,
        'items': items,
        'status_choices': status_choices,
        'title': f'Деталі заявки №{work_request.pk}'
    }
    return render(request, 'oids/lists/work_request_detail.html', context)

@transaction.atomic
def plan_trip_view(request):
    if request.method == 'POST':
        form = TripForm(request.POST)
        if form.is_valid():
            trip = form.save(commit=False) 
            # Тут можна додати будь-яку логіку перед першим збереженням Trip, якщо потрібно
            trip.save() # Перше збереження для отримання trip.id
            form.save_m2m() # Зберігаємо ManyToMany зв'язки (units, oids, work_requests)

            # --- Логіка встановлення дедлайнів ПІСЛЯ збереження M2M ---
            if trip.end_date:
                print(f"PLAN_TRIP_VIEW: Trip ID {trip.pk} saved with end_date: {trip.end_date}. Calculating deadlines.")
                
                linked_work_requests = trip.work_requests.all()
                linked_oids_direct = trip.oids.all()
                
                print(f"PLAN_TRIP_VIEW: Trip linked to WorkRequest IDs: {[wr.id for wr in linked_work_requests]}")
                print(f"PLAN_TRIP_VIEW: Trip directly linked to OID IDs: {[o.id for o in linked_oids_direct]}")

                if linked_work_requests.exists():
                    items_to_process = WorkRequestItem.objects.filter(
                        request__in=linked_work_requests,
                        oid__in=linked_oids_direct 
                        # Можна додати: doc_processing_deadline__isnull=True
                    ).select_related('oid', 'request__unit')

                    print(f"PLAN_TRIP_VIEW: Found {items_to_process.count()} WRI to update for trip {trip.id}")
                    
                    start_counting_from_date = trip.end_date

                    for item in items_to_process:
                        days_for_processing = 0
                        if item.work_type == WorkTypeChoices.IK: days_for_processing = 10 # налаштувати кількість днів для розрахунку часу на опрацювання   
                        elif item.work_type == WorkTypeChoices.ATTESTATION: days_for_processing = 15 # налаштувати кількість днів для розрахунку часу на опрацювання 
                        elif item.work_type == WorkTypeChoices.PLAND_ATTESTATION: days_for_processing = 15 # налаштувати кількість днів для розрахунку часу на опрацювання 
                        
                        if days_for_processing > 0:
                           new_deadline = add_working_days(start_counting_from_date, days_for_processing)
                           if item.doc_processing_deadline != new_deadline or item.deadline_trigger_trip != trip: # Додано перевірку trip
                               item.doc_processing_deadline = new_deadline
                               item.deadline_trigger_trip = trip # <--- Встановлюємо зв'язок
                               item.save(update_fields=['doc_processing_deadline', 'deadline_trigger_trip', 'updated_at'])
                               print(f"PLAN_TRIP_VIEW: WRI ID {item.id} (OID: {item.oid.cipher}) deadline -> {item.doc_processing_deadline}")
                else:
                    print(f"PLAN_TRIP_VIEW: No WorkRequests linked to Trip ID {trip.pk} to calculate item deadlines.")
            else:
                print(f"PLAN_TRIP_VIEW: Trip ID {trip.pk} has no end_date. Deadline calculation skipped.")
            # --- Кінець логіки дедлайнів ---


            messages.success(request, f'Відрядження заплановано успішно (ID: {trip.id}).')
            # Оновлюємо статус пов'язаних заявок на "В роботі" (ця логіка у вас вже була)
            for work_request in linked_work_requests: # Використовуємо вже отриманий queryset
                if work_request.status == WorkRequestStatusChoices.PENDING:
                    work_request.status = WorkRequestStatusChoices.IN_PROGRESS
                    work_request.save(update_fields=['status', 'updated_at'])
                    WorkRequestItem.objects.filter(request=work_request, status=WorkRequestStatusChoices.PENDING)\
                                           .update(status=WorkRequestStatusChoices.IN_PROGRESS)

            return redirect('oids:list_trips') # Або main_dashboard
        else:
            messages.error(request, 'Будь ласка, виправте помилки у формі.')
            print(f"PLAN_TRIP_VIEW: Form errors: {form.errors.as_json()}")
    else: # GET request
        form = TripForm()
    
    context = {
        'form': form, 
        'page_title': 'Запланувати відрядження'
    }
    return render(request, 'oids/forms/plan_trip_form.html', context)

def get_last_document_expiration_date(oid_instance, document_name_keyword, work_type_choice=None):
    try:
        doc_type_filters = Q(name__icontains=document_name_keyword)
        if work_type_choice:
            doc_type_filters &= Q(work_type=work_type_choice)
        
        relevant_doc_type_qs = DocumentType.objects.filter(doc_type_filters)
        if not relevant_doc_type_qs.exists():
            return None
        relevant_doc_type = relevant_doc_type_qs.first()

        last_document = Document.objects.filter(
            oid=oid_instance, # Використовуємо переданий екземпляр OID
            document_type=relevant_doc_type,
            expiration_date__isnull=False
        ).order_by('-work_date', '-doc_process_date').first()
        
        return last_document.expiration_date if last_document else None
    except Exception as e:
        print(f"Помилка get_last_document_expiration_date для ОІД {oid_instance.cipher if oid_instance else 'N/A'} ({document_name_keyword}): {e}")
        return None

@login_required 
def add_document_processing_view(request, oid_id=None, work_request_item_id=None):
    selected_oid_instance = None
    selected_wri_instance = None # WorkRequestItem instance
    initial_main_form_data = {}

    if oid_id:
        selected_oid_instance = get_object_or_404(OID, pk=oid_id)
        initial_main_form_data['unit'] = selected_oid_instance.unit
        initial_main_form_data['oid'] = selected_oid_instance
    
    if work_request_item_id:
        selected_wri_instance = get_object_or_404(WorkRequestItem, pk=work_request_item_id)
        initial_main_form_data['work_request_item'] = selected_wri_instance
        if not selected_oid_instance: # Якщо ОІД не передано, беремо з WorkRequestItem
            selected_oid_instance = selected_wri_instance.oid
            initial_main_form_data['unit'] = selected_oid_instance.unit # Оновлюємо unit, якщо потрібно
            initial_main_form_data['oid'] = selected_oid_instance

    if request.method == 'POST':
        main_form = DocumentProcessingMainForm(request.POST, prefix='main')
        # Передаємо initial_oid та initial_work_request_item для правильного встановлення queryset у __init__
        # formset = DocumentItemFormSet(request.POST, request.FILES, prefix='docs', form_kwargs={'initial_oid': selected_oid_instance, 'initial_wri': selected_wri_instance})
        # form_kwargs тут не спрацює для formset_factory напряму, логіку queryset для DocumentItemForm.document_type краще робити в JS
        formset = DocumentItemFormSet(request.POST, request.FILES, prefix='docs')


        if main_form.is_valid() and formset.is_valid():
            # Отримуємо дані з головної форми
            oid_instance = main_form.cleaned_data['oid']
            work_request_item_instance = main_form.cleaned_data.get('work_request_item') # Може бути None
            doc_process_date_from_main = main_form.cleaned_data['doc_process_date']
            work_date_from_main = main_form.cleaned_data['work_date']
            author_instance = main_form.cleaned_data.get('author')

            saved_docs_count = 0
            for item_form in formset:
                if item_form.is_valid() and item_form.has_changed(): # Обробляємо тільки валідні та змінені форми
                    document_instance = item_form.save(commit=False)
                    document_instance.oid = oid_instance
                    document_instance.work_request_item = work_request_item_instance
                    document_instance.doc_process_date = doc_process_date_from_main
                    document_instance.work_date = work_date_from_main
                    document_instance.author = author_instance
                    
                    # Логіка обчислення expiration_date вже є в моделі Document.save()
                    document_instance.save() # Це викличе Document.save() з усіма обчисленнями
                    saved_docs_count += 1
            
            if saved_docs_count > 0:
                messages.success(request, f'{saved_docs_count} документ(ів) успішно додано до ОІД "{oid_instance.cipher}".')
                # Оновлення статусів (як було, але тепер контекст може бути іншим)
                if work_request_item_instance:
                    # Логіка оновлення статусу WorkRequestItem (можливо, перевірка всіх доданих документів)
                    # work_request_item_instance.status = ...
                    # work_request_item_instance.save() -> це викличе оновлення статусу головної заявки
                    pass 
                if oid_instance:
                    # Логіка оновлення статусу ОІД
                    pass
                return redirect('oids:oid_detail_view_name', oid_id=oid_instance.id)
            else:
                messages.info(request, "Не було додано жодного нового документа.")
                # Залишаємося на сторінці або перенаправляємо
        else:
            messages.error(request, "Будь ласка, виправте помилки у формі.")
            if not main_form.is_valid():
                print("Main form errors:", main_form.errors.as_json())
            if not formset.is_valid():
                print("Formset errors:", formset.errors) # formset.errors - це список словників
                for i, form_errors in enumerate(formset.errors):
                    if form_errors:
                        print(f"Errors in form {i}: {form_errors}")


    else: # GET request
        main_form = DocumentProcessingMainForm(initial=initial_main_form_data, prefix='main', 
                                               initial_oid=selected_oid_instance, 
                                               initial_work_request_item=selected_wri_instance)
        formset = DocumentItemFormSet(prefix='docs')

    context = {
        'main_form': main_form,
        'formset': formset,
        'page_title': 'Додати опрацювання документів',
        'selected_oid': selected_oid_instance # Для відображення в заголовку
    }
    return render(request, 'oids/forms/add_document_processing_form.html', context)

@login_required 
def update_oid_status_view(request, oid_id_from_url=None):
    initial_unit = None
    initial_oid = None

    if oid_id_from_url: # Якщо переходимо на сторінку для конкретного ОІД
        initial_oid = get_object_or_404(OID, pk=oid_id_from_url)
        initial_unit = initial_oid.unit

    if request.method == 'POST':
        form = OIDStatusUpdateForm(request.POST, initial_unit_id=initial_unit.id if initial_unit else None, 
                                   initial_oid_id=initial_oid.id if initial_oid else None)
        if form.is_valid():
            oid_to_update = form.cleaned_data['oid']
            new_status = form.cleaned_data['new_status']
            reason_for_change = form.cleaned_data['reason']
            changed_by_person = form.cleaned_data.get('changed_by') # Може бути None
            doc_number = form.cleaned_data.get('initiating_document_number')
            doc_date = form.cleaned_data.get('initiating_document_date')

            old_status_val = oid_to_update.status 
            
            # Створюємо запис в історії
            OIDStatusChange.objects.create(
                oid=oid_to_update,
                old_status=old_status_val, # Зберігаємо старий статус
                new_status=new_status,
                reason=reason_for_change,
                changed_by=changed_by_person, # TODO: Замінити на request.user, коли буде автентифікація
                # Якщо потрібно пов'язати з документом, тут потрібна логіка пошуку/створення Document
                # initiating_document= ... 
            )
            
            # Оновлюємо статус самого ОІД
            oid_to_update.status = new_status
            oid_to_update.save(update_fields=['status', 'updated_at']) # Оновлюємо тільки потрібні поля

            messages.success(request, f"Статус для ОІД '{oid_to_update.cipher}' успішно змінено на '{oid_to_update.get_status_display()}'.")
            return redirect('oids:oid_detail_view_name', oid_id=oid_to_update.id)
        else:
            messages.error(request, "Будь ласка, виправте помилки у формі.")
            # Якщо форма не валідна, selected_oid для заголовка може бути втрачений, 
            # потрібно його отримати знову, якщо можливо
            submitted_oid_id = request.POST.get('oid')
            if submitted_oid_id and submitted_oid_id.isdigit():
                 try:
                    selected_oid_instance = OID.objects.get(pk=int(submitted_oid_id))
                 except OID.DoesNotExist:
                    selected_oid_instance = None
            else:
                selected_oid_instance = initial_oid # Якщо був GET параметр
    else: # GET request
        form = OIDStatusUpdateForm(initial_unit_id=initial_unit.id if initial_unit else None, 
                                   initial_oid_id=initial_oid.id if initial_oid else None)
        selected_oid_instance = initial_oid


    context = {
        'form': form,
        'page_title': 'Зміна статусу Об\'єкта Інформаційної Діяльності (ОІД)',
        'selected_oid_for_title': selected_oid_instance, # Для відображення в заголовку
        'current_oid_status_display': selected_oid_instance.get_status_display() if selected_oid_instance else None,
    }
    return render(request, 'oids/forms/update_oid_status_form.html', context)

@login_required 
@transaction.atomic
def send_attestation_for_registration_view(request):
    """
    Формування відправки Актів Атестації на реєстрацію в ДССЗЗІ
    """
    if request.method == 'POST':
        form = AttestationRegistrationSendForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                # Зберігаємо форму (вона сама створює AttestationRegistration та зв'язує документи)
                attestation_registration = form.save()
                
                # Отримуємо дані для повідомлення
                sent_documents = attestation_registration.sent_documents.all()
                units = attestation_registration.units.all()
                
                # Групуємо документи по ВЧ для зручності
                units_dict = {}
                for doc in sent_documents:
                    unit_code = doc.oid.unit.code if doc.oid.unit else 'Без ВЧ'
                    if unit_code not in units_dict:
                        units_dict[unit_code] = []
                    
                    units_dict[unit_code].append({
                        'cipher': doc.oid.cipher,
                        'doc_number': doc.document_number,
                        'doc_date': doc.doc_process_date,
                        'doc_type': doc.document_type.name if doc.document_type else 'Невідомий тип'
                    })
                
                # Формуємо детальне повідомлення
                success_message = (
                    f"✅ <strong>Відправку на реєстрацію успішно сформовано!</strong><br>"
                    f"📋 Вихідний лист: <strong>№{attestation_registration.outgoing_letter_number}</strong> "
                    f"від <strong>{attestation_registration.outgoing_letter_date.strftime('%d.%m.%Y')}</strong><br>"
                    f"📄 Актів атестації: <strong>{sent_documents.count()}</strong><br>"
                    f"🏢 Військових частин: <strong>{units.count()}</strong><br>"
                )
                
                # Додаємо інформацію про відправника, якщо є
                if attestation_registration.sent_by:
                    success_message += f"👤 Відправник: <strong>{attestation_registration.sent_by.full_name}</strong><br>"
                
                success_message += "<br><strong>Деталі по військових частинах:</strong><br>"
                
                # Додаємо інформацію згруповану по ВЧ
                for unit_code, docs_list in units_dict.items():
                    success_message += f"<br><strong>📍 ВЧ {unit_code}</strong> ({len(docs_list)} актів):<br>"
                    for doc_info in docs_list:
                        success_message += (
                            f"&nbsp;&nbsp;&nbsp;• ОІД: <strong>{doc_info['cipher']}</strong><br>"
                            f"&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;{doc_info['doc_type']}: №<strong>{doc_info['doc_number']}</strong> "
                        )
                        if doc_info['doc_date']:
                            success_message += f"від {doc_info['doc_date'].strftime('%d.%m.%Y')}"
                        success_message += "<br>"
                
                messages.success(request, mark_safe(success_message))
                
                # Додаткове інформаційне повідомлення
                messages.info(
                    request,
                    mark_safe(
                        f"ℹ️ Наступний крок: після отримання відповіді від ДССЗЗІ, "
                        f"<a href=\"{reverse('oids:record_attestation_response_for_registration', args=[attestation_registration.id])}\">внесіть реєстраційні номери</a>. "
                        f"Або переглянте <a href=\"{reverse('oids:list_attestation_registrations')}\">всі відправки</a>."
                    )
                )
                
                return redirect('oids:list_attestation_registrations')
                
            except Exception as e:
                # Обробка помилок
                messages.error(
                    request,
                    mark_safe(
                        f"❌ <strong>Помилка при створенні відправки:</strong><br>"
                        f"{str(e)}<br><br>"
                        f"Зміни не були збережені."
                    )
                )
                print(f"SendAttestationForm save error: {str(e)}")
                return redirect('oids:send_attestation_for_registration')
        
        else:
            # Якщо форма не валідна - показуємо помилки
            error_messages = []
            
            # Загальні помилки форми
            if form.non_field_errors():
                error_messages.append("<strong>Загальні помилки:</strong>")
                for error in form.non_field_errors():
                    error_messages.append(f"• {error}")
            
            # Помилки конкретних полів
            for field_name, errors in form.errors.items():
                if field_name == '__all__':
                    continue
                
                field_label = form.fields[field_name].label or field_name
                
                # Спеціальна обробка для різних полів
                if field_name == 'selected_units':
                    error_messages.append("<br><strong>Крок 1 - Військові частини:</strong>")
                elif field_name == 'selected_oids':
                    error_messages.append("<br><strong>Крок 2 - ОІДи:</strong>")
                elif field_name == 'attestation_acts_to_send':
                    error_messages.append("<br><strong>Крок 3 - Акти атестації:</strong>")
                
                for error in errors:
                    error_messages.append(f"• {field_label}: {error}")
            
            messages.error(
                request,
                mark_safe(
                    "❌ <strong>Будь ласка, виправте помилки у формі:</strong><br>" + 
                    "<br>".join(error_messages)
                )
            )
            
            # Додаємо помилки в консоль для дебагу
            print(f"SendAttestationForm errors: {form.errors.as_json()}")
    
    else:  # GET request
        form = AttestationRegistrationSendForm()

    context = {
        'form': form,
        'page_title': 'Сформувати відправку Актів Атестації на реєстрацію'
    }
    return render(request, 'oids/forms/send_attestation_form.html', context)

# ДОДАТКОВА ФУНКЦІЯ: Статистика про відправку (опціонально)
def get_attestation_registration_stats(attestation_registration):
    """
    Повертає детальну статистику про відправку
    Можна використовувати для додаткових повідомлень
    """
    stats = {
        'total_documents': attestation_registration.sent_documents.count(),
        'total_units': attestation_registration.units.count(),
        'total_oids': attestation_registration.oids.count(),
        'documents_by_unit': {},
        'documents_by_type': {},
    }
    
    # Групування по ВЧ
    for doc in attestation_registration.sent_documents.all():
        unit_code = doc.oid.unit.code if doc.oid.unit else 'Без ВЧ'
        if unit_code not in stats['documents_by_unit']:
            stats['documents_by_unit'][unit_code] = 0
        stats['documents_by_unit'][unit_code] += 1
        
        # Групування по типу документу
        doc_type = doc.document_type.name if doc.document_type else 'Невідомий'
        if doc_type not in stats['documents_by_type']:
            stats['documents_by_type'][doc_type] = 0
        stats['documents_by_type'][doc_type] += 1
    
    return stats

@login_required
def send_document_for_registration_view(request, pk):
    document = get_object_or_404(Document, pk=pk)
    
    # Перевірка, чи можна виконати цю дію
    if document.processing_status != DocumentProcessingStatusChoices.DRAFT:
        messages.warning(request, "Цей документ вже було відправлено або опрацьовано.")
        return redirect('oids:oid_detail', pk=document.oid.pk)

    if request.method == 'POST':
        form = SendForRegistrationForm(request.POST)
        if form.is_valid():
            # Основна логіка: змінюємо статус і зберігаємо
            document.processing_status = DocumentProcessingStatusChoices.SENT_FOR_REGISTRATION
            document.save(update_fields=['processing_status', 'updated_at'])
            
            # Сигнал `post_save` автоматично оновить крок процесу в цей момент
            
            messages.success(request, f"Документ '{document}' було успішно відправлено на реєстрацію.")
            return redirect('oids:oid_detail', pk=document.oid.pk)
    else:
        form = SendForRegistrationForm()
        
    context = {
        'form': form,
        'document': document,
        'page_title': f'Підтвердження: Відправити на реєстрацію {document}'
    }
    return render(request, 'oids/action_confirm_form.html', context)

@login_required
def send_azr_for_registration_view(request):
    """
    Обробляє сторінку "Відправка АЗР" з динамічним створенням полів.
    """
    if request.method == 'POST':
        submission_form = AzrSubmissionForm(request.POST)
        item_formset = AzrItemFormSet(request.POST)

        if submission_form.is_valid() and item_formset.is_valid():
            try:
                # 1. Створюємо "конверт" - запис про відправку
                registration_request = WorkCompletionRegistration.objects.create(
                    outgoing_letter_number=submission_form.cleaned_data['outgoing_letter_number'],
                    outgoing_letter_date=submission_form.cleaned_data['outgoing_letter_date'],
                    note=submission_form.cleaned_data['note'],
                    # created_by=request.user.person # Якщо потрібно
                )

                # Перевіряємо наявність типу документу
                try:
                    azr_doc_type = DocumentType.objects.get(name__icontains="Акт завершення")
                except DocumentType.DoesNotExist:
                    messages.error(
                        request, 
                        "❌ Критична помилка: Тип документу 'Акт завершення робіт' не знайдено в системі."
                    )
                    return redirect('oids:main_dashboard')

                docs_created = []
                oids_for_m2m = []
                oids_info = []  # Для детального повідомлення
                
                # 2. Проходимо по даних з формсету і створюємо документи
                with transaction.atomic():
                    for form in item_formset:
                        oid_id = form.cleaned_data.get('oid')
                        if not oid_id:
                            continue
                        
                        oid_instance = OID.objects.get(pk=oid_id)
                        
                        # Створюємо документ стандартним способом (викличеться save())
                        doc = Document(
                            oid=oid_instance,
                            document_type=azr_doc_type,
                            document_number=form.cleaned_data.get('prepared_number'),
                            doc_process_date=form.cleaned_data.get('prepared_date'),
                            work_date=form.cleaned_data.get('prepared_date'),
                            wcr_submission=registration_request,
                            # author=request.user.person
                        )
                        doc.save()  # Це запустить логіку оновлення статусів
                        
                        docs_created.append(doc)
                        oids_for_m2m.append(oid_instance)
                        
                        # Збираємо інформацію для повідомлення
                        oids_info.append({
                            'cipher': oid_instance.cipher,
                            'unit': oid_instance.unit.code if oid_instance.unit else 'Без ВЧ',
                            'doc_number': doc.document_number
                        })
                
                # 3. Додаємо ОІДи до m2m поля "конверта"
                registration_request.oids.set(oids_for_m2m)
                
                # 4. Формуємо детальне повідомлення про успіх
                success_message = (
                    f"✅ <strong>Успішно відправлено на реєстрацію!</strong><br>"
                    f"📋 Вихідний лист: <strong>№{registration_request.outgoing_letter_number}</strong> "
                    f"від <strong>{registration_request.outgoing_letter_date.strftime('%d.%m.%Y')}</strong><br>"
                    f"📄 Створено актів завершення робіт: <strong>{len(docs_created)}</strong><br>"
                    f"<br><strong>Деталі:</strong><br>"
                )
                
                # Додаємо список ОІДів
                for info in oids_info:
                    success_message += (
                        f"• ВЧ <strong>{info['unit']}</strong> | "
                        f"ОІД <strong>{info['cipher']}</strong> | "
                        f"АЗР №<strong>{info['doc_number']}</strong><br>"
                    )
                
                messages.success(request, mark_safe(success_message))
                
                # Додаткове інформаційне повідомлення
                messages.info(
                    request,
                    f"ℹ️ Ви можете переглянути статус відправки в <a href=\"{reverse('oids:list_azr_registrations')}\">списку відправок АЗР</a>",
                    extra_tags='safe'
                )
                
                return redirect('oids:list_azr_registrations')
                
            except Exception as e:
                # Обробка помилок
                messages.error(
                    request,
                    f"❌ <strong>Помилка при створенні відправки:</strong><br>{str(e)}",
                    extra_tags='safe'
                )
                return redirect('oids:send_azr_for_registration')
        
        else:
            # Якщо форма або формсет не валідні
            error_messages = []
            
            if not submission_form.is_valid():
                error_messages.append("<strong>Помилки в даних супровідного листа:</strong>")
                for field, errors in submission_form.errors.items():
                    field_label = submission_form.fields[field].label or field
                    for error in errors:
                        error_messages.append(f"• {field_label}: {error}")
            
            if not item_formset.is_valid():
                error_messages.append("<br><strong>Помилки в даних актів:</strong>")
                for i, form_errors in enumerate(item_formset.errors):
                    if form_errors:
                        error_messages.append(f"<br><strong>Акт #{i+1}:</strong>")
                        for field, errors in form_errors.items():
                            for error in errors:
                                error_messages.append(f"• {field}: {error}")
            
            messages.error(
                request,
                mark_safe("❌ <strong>Виправте помилки у формі:</strong><br>" + "<br>".join(error_messages))
            )
    
    else:
        submission_form = AzrSubmissionForm()
        item_formset = AzrItemFormSet()

    context = {
        'submission_form': submission_form,
        'item_formset': item_formset,
        'unit_selector': forms.ModelMultipleChoiceField(
            queryset=Unit.objects.order_by('name').filter(is_active=True),
            required=False
        ),
        'page_title': 'Відправка Актів Завершення Робіт на реєстрацію'
    }
    return render(request, 'oids/forms/send_azr_for_registration.html', context)

@login_required
def record_azr_response_view(request, registration_id):
    """
    Обробляє сторінку "Внесення відповіді по 'Акт завершення робіт' ".
    """
    registration_request = get_object_or_404(WorkCompletionRegistration, pk=registration_id)
    # Отримуємо queryset тільки тих АЗР, які були в цій конкретній відправці
    queryset_for_formset = registration_request.submitted_documents.all()

    if request.method == 'POST':
        response_form = WorkCompletionResponseForm(request.POST)
        formset = AzrUpdateFormSet(request.POST, queryset=queryset_for_formset)

        if response_form.is_valid() and formset.is_valid():
            # Створюємо запис про сам факт отримання відповіді
            response_instance = response_form.save(commit=False)
            response_instance.registration_request = registration_request
            # response_instance.created_by = request.user.person
            response_instance.save()
            
            # Зберігаємо оновлені дані для кожного АЗР (реєстраційні номери/дати).
            # Метод .save() для формсету викличе .save() для кожного екземпляра
            # моделі Document, що, в свою чергу, запустить нашу логіку
            # оновлення статусів OID та WorkRequestItem, якщо вона там є.
            formset.save()

            messages.success(request, "Відповідь по 'Акт завершення робіт' успішно внесено.")
            return redirect('oids:list_azr_documents')
    else:
        response_form = WorkCompletionResponseForm()
        formset = AzrUpdateFormSet(queryset=queryset_for_formset)

    context = {
        'response_form': response_form,
        'formset': formset,
        'registration_request': registration_request,
        'page_title': f'Внесення відповіді на лист № {registration_request.outgoing_letter_number}'
    }
    return render(request, 'oids/forms/record_azr_response.html', context)


# view declaration


@login_required
@transaction.atomic
def send_declaration_for_registration_view(request):
    """
    ОНОВЛЕНА ВЕРСІЯ: Створює декларації та відображає детальні повідомлення.
    """
    if request.method == 'POST':
        submission_form = DeclarationSubmissionForm(request.POST)
        item_formset = DeclarationItemFormSet(request.POST, prefix='items')

        if submission_form.is_valid() and item_formset.is_valid():
            try:
                # 1. Створюємо "конверт"
                submission = submission_form.save(commit=False)
                # submission.created_by = request.user.person # Якщо потрібно
                submission.save()

                declarations_to_link = []
                declarations_info = []  # Для детального повідомлення
                units_dict = {}  # Групування по ВЧ
                
                # 2. Проходимо по всіх формах з формсету
                for form in item_formset:
                    # 3. Створюємо ДСК ЕОТ
                    unit = form.cleaned_data.get('unit')
                    cipher = form.cleaned_data.get('cipher')
                    
                    dsk_eot = DskEot.objects.create(
                        unit=unit,
                        cipher=cipher,
                        serial_number=form.cleaned_data.get('serial_number'),
                        inventory_number=form.cleaned_data.get('inventory_number'),
                        room=form.cleaned_data.get('room'),
                    )
                    
                    # 4. Створюємо відповідну Декларацію
                    declaration = Declaration.objects.create(
                        dsk_eot=dsk_eot,
                        prepared_number=form.cleaned_data.get('prepared_number'),
                        prepared_date=form.cleaned_data.get('prepared_date'),
                    )
                    declarations_to_link.append(declaration)
                    
                    # Збираємо інформацію для повідомлення
                    unit_code = unit.code if unit else 'Без ВЧ'
                    if unit_code not in units_dict:
                        units_dict[unit_code] = []
                    
                    units_dict[unit_code].append({
                        'cipher': cipher,
                        'prepared_number': declaration.prepared_number,
                        'prepared_date': declaration.prepared_date,
                        'serial_number': dsk_eot.serial_number or '-',
                        'room': dsk_eot.room
                    })

                # 5. Прив'язуємо всі створені декларації до нашої відправки
                if declarations_to_link:
                    submission.declarations.set(declarations_to_link)

                # 6. Формуємо детальне повідомлення про успіх
                success_message = (
                    f"✅ <strong>Успішно відправлено на реєстрацію!</strong><br>"
                    f"📋 Вихідний лист: <strong>№{submission.outgoing_letter_number}</strong> "
                    f"від <strong>{submission.outgoing_letter_date.strftime('%d.%m.%Y')}</strong><br>"
                    f"📄 Створено декларацій відповідності: <strong>{len(declarations_to_link)}</strong><br>"
                    f"🏢 Військових частин: <strong>{len(units_dict)}</strong><br>"
                    f"<br><strong>Деталі по військових частинах:</strong><br>"
                )
                
                # Додаємо інформацію згруповану по ВЧ
                for unit_code, declarations_list in units_dict.items():
                    success_message += f"<br><strong>📍 ВЧ {unit_code}</strong> ({len(declarations_list)} декл.):<br>"
                    for decl_info in declarations_list:
                        success_message += (
                            f"&nbsp;&nbsp;&nbsp;• Шифр: <strong>{decl_info['cipher']}</strong> | "
                            f"Підг. №<strong>{decl_info['prepared_number']}</strong> "
                            f"від {decl_info['prepared_date'].strftime('%d.%m.%Y')}<br>"
                            f"&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;Серійний №: {decl_info['serial_number']} | "
                            f"Приміщення: {decl_info['room']}<br>"
                        )
                
                messages.success(request, mark_safe(success_message))
                
                # Додаткове інформаційне повідомлення
                messages.info(
                    request,
                    mark_safe(
                        f"ℹ️ Ви можете переглянути статус відправки в "
                        f"<a href=\"{reverse('oids:list_declaration_registrations')}\">списку відправок Декларацій</a>"
                    )
                )
                
                return redirect('oids:list_declaration_registrations')
                
            except Exception as e:
                # Обробка помилок
                messages.error(
                    request,
                    mark_safe(
                        f"❌ <strong>Помилка при створенні відправки:</strong><br>"
                        f"{str(e)}<br><br>"
                        f"Зміни не були збережені."
                    )
                )
                return redirect('oids:send_declaration_for_registration')
        
        else:
            # Якщо форма або формсет не валідні
            error_messages = []
            
            if not submission_form.is_valid():
                error_messages.append("<strong>Помилки в даних супровідного листа:</strong>")
                for field, errors in submission_form.errors.items():
                    field_label = submission_form.fields[field].label or field
                    for error in errors:
                        error_messages.append(f"• {field_label}: {error}")
            
            if not item_formset.is_valid():
                error_messages.append("<br><strong>Помилки в даних декларацій:</strong>")
                for i, form_errors in enumerate(item_formset.errors):
                    if form_errors:
                        error_messages.append(f"<br><strong>Декларація #{i+1}:</strong>")
                        for field, errors in form_errors.items():
                            for error in errors:
                                error_messages.append(f"• {field}: {error}")
                
                # Показуємо також non_form_errors якщо є
                if item_formset.non_form_errors():
                    error_messages.append("<br><strong>Загальні помилки формсету:</strong>")
                    for error in item_formset.non_form_errors():
                        error_messages.append(f"• {error}")
            
            messages.error(
                request,
                mark_safe(
                    "❌ <strong>Виправте помилки у формі:</strong><br>" + 
                    "<br>".join(error_messages)
                )
            )
    
    else:
        submission_form = DeclarationSubmissionForm()
        item_formset = DeclarationItemFormSet(prefix='items')

    context = {
        'submission_form': submission_form,
        'item_formset': item_formset,
        'all_units': Unit.objects.all().order_by('code'),
        'page_title': 'Відправка Декларацій відповідності на реєстрацію'
    }
    return render(request, 'oids/forms/send_declaration_for_registration.html', context)


@login_required
def list_declaration_registrations_view(request):
    """
    Відображає список усіх ВІДПРАВОК Декларацій на реєстрацію.
    """
    # prefetch_related('declarations') - оптимізація, щоб уникнути зайвих запитів до БД
    all_submissions = DeclarationRegistration.objects.prefetch_related('declarations').order_by('-outgoing_letter_date')
    
    page_obj = get_paginated_page(all_submissions, request)    

    context = {
        'page_title': 'Декларації відправлені на реєстрацію до ДССЗЗІ',
        'submissions': page_obj,
    }
    return render(request, 'oids/lists/declaration_registration_list.html', context)



@login_required
def declaration_list_view(request):
    """
    Відображає список всіх Декларацій відповідності з фільтрацією та сортуванням.
    """
    queryset = Declaration.objects.select_related('dsk_eot__unit').prefetch_related('registrations').all()
    
    # --- Фільтрація ---
    filter_form = DeclarationFilterForm(request.GET or None)
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('unit'):
            queryset = queryset.filter(dsk_eot__unit__in=filter_form.cleaned_data['unit'])
        if filter_form.cleaned_data.get('dsk_eot'):
            queryset = queryset.filter(dsk_eot__in=filter_form.cleaned_data['dsk_eot'])
        if filter_form.cleaned_data.get('prepared_number'):
            queryset = queryset.filter(prepared_number__icontains=filter_form.cleaned_data['prepared_number'])
        if filter_form.cleaned_data.get('registered_number'):
            queryset = queryset.filter(registered_number__icontains=filter_form.cleaned_data['registered_number'])
        if filter_form.cleaned_data.get('date_from'):
            queryset = queryset.filter(prepared_date__gte=filter_form.cleaned_data['date_from'])
        if filter_form.cleaned_data.get('date_to'):
            queryset = queryset.filter(prepared_date__lte=filter_form.cleaned_data['date_to'])

    # --- Сортування ---
    sort_by = request.GET.get('sort', '-prepared_date')
    valid_sort_fields = [
        'dsk_eot__unit__code',
        'dsk_eot__cipher',
        'prepared_number',
        'prepared_date',
        'registered_number',
        'registered_date'
        ]
    if sort_by.lstrip('-') in valid_sort_fields:
        queryset = queryset.order_by(sort_by)
    
	
    # --- Експорт в Excel ---
    if 'export' in request.GET:
        columns = {
            'dsk_eot__unit__code': 'ВЧ',
            'dsk_eot__cipher': 'Шифр ДСК ЕОТ',
            'dsk_eot__serial_number': 'Серійний №',
            'dsk_eot__inventory_number': 'Інв. №',
            'dsk_eot__room': 'Приміщення',
            'prepared_number': 'Підготовлений №',
            'prepared_date': 'Дата підготовки',
            'registered_number': 'Зареєстрований №',
            'registered_date': 'Дата реєстрації',
            'get_status_display': 'Статус',
            'get_submission_info': 'Інформація про відправку',
        }
        return export_to_excel(
            queryset,
            columns,
            filename='declarations.xlsx',
            include_row_numbers=True 
            )
    # --- Пагінація ---
    page_obj = get_paginated_page(queryset, request)
    
    query_params = request.GET.copy()
    if 'sort' in query_params:
        del query_params['sort']
    if 'page' in query_params:
        del query_params['page']

    context = {
        'page_title': 'Реєстр Декларацій відповідності',
        'declarations': page_obj,
        'filter_form': filter_form,
        'current_sort': sort_by,
        'sort_url_part': query_params.urlencode(),
    }
    return render(request, 'oids/lists/declaration_list.html', context)

# oids/views.py

@login_required
def record_declaration_response_view(request, submission_id):
    """ 
    Обробляє сторінку "Внесення відповіді по Деклараціях".
    ОНОВЛЕНА ВЕРСІЯ: Правильно оновлює існуючий запис.
    """
    # Отримуємо існуючий запис про відправку, який ми будемо оновлювати
    submission = get_object_or_404(DeclarationRegistration, pk=submission_id)
    queryset_for_formset = submission.declarations.select_related('dsk_eot__unit').all()

    if request.method == 'POST':
        # Ми оновлюємо існуючий об'єкт, тому передаємо `instance=submission`
        response_form = DeclarationResponseForm(request.POST, instance=submission)
        formset = DeclarationUpdateFormSet(request.POST, queryset=queryset_for_formset)

        if response_form.is_valid() and formset.is_valid():
            # 1. Зберігаємо дані про відповідь В ТОЙ САМИЙ об'єкт відправки
            response_instance = response_form.save(commit=False)
            response_instance.response_at = timezone.now()
            # response_instance.response_by = request.user.person
            response_instance.save()
            
            # 2. Зберігаємо реєстраційні дані для кожної Декларації
            formset.save()

            messages.success(request, "Відповідь по Деклараціях успішно внесено.")
            return redirect('oids:list_declaration_registrations')
    else:
        # При першому завантаженні також передаємо `instance`,
        # щоб форма могла показати вже існуючі дані, якщо вони є.
        response_form = DeclarationResponseForm(instance=submission)
        formset = DeclarationUpdateFormSet(queryset=queryset_for_formset)

    context = {
        'response_form': response_form,
        'formset': formset,
        'submission': submission,
        'page_title': f'Внесення відповіді на лист № {submission.outgoing_letter_number}'
    }
    return render(request, 'oids/forms/record_declaration_response.html', context)


@login_required 
def technical_task_create_view(request):
    initial_data_for_form = {}
    # Якщо передано Unit або OID через GET, використовуємо їх для початкового заповнення
    unit_id_get = request.GET.get('unit_id')
    oid_id_get = request.GET.get('oid_id')

    if unit_id_get:
        initial_data_for_form['unit'] = unit_id_get
    if oid_id_get:
        initial_data_for_form['oid'] = oid_id_get
        if not unit_id_get: # Якщо є ОІД, але немає ВЧ, спробуємо її отримати
            try:
                oid_instance = OID.objects.get(pk=oid_id_get)
                initial_data_for_form['unit'] = oid_instance.unit_id
            except OID.DoesNotExist:
                pass

    if request.method == 'POST':
        form = TechnicalTaskCreateForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.review_result = DocumentReviewResultChoices.READ # Встановлюємо статус "Опрацювати"
            # task.created_by = request.user # Якщо є автентифікація і поле для автора створення
            task.save()
            messages.success(request, f"Технічне завдання №{task.input_number} для ОІД '{task.oid.cipher}' успішно внесено зі статусом 'Опрацювати'.")
            return redirect('oids:list_technical_tasks') # Або на деталі ТЗ, або на список
    else:
        form = TechnicalTaskCreateForm(initial=initial_data_for_form, 
                                       initial_unit_id=initial_data_for_form.get('unit'),
                                       initial_oid_id=initial_data_for_form.get('oid'))

    context = {
        'form': form,
        'page_title': 'Внесення нового Технічного Завдання'
    }
    return render(request, 'oids/forms/technical_task_create_form.html', context)

@login_required 
def technical_task_process_view(request, task_id=None): # Може приймати ID ТЗ з URL
    # Якщо task_id передано, це форма для конкретного ТЗ.
    # Якщо ні, то перше поле форми - вибір ТЗ.
    
    technical_task_instance = None
    if task_id:
        technical_task_instance = get_object_or_404(TechnicalTask, pk=task_id, review_result=DocumentReviewResultChoices.READ)
    
    if request.method == 'POST':
        # Якщо ми працюємо з конкретним ТЗ, передаємо його instance у форму
        # Якщо ТЗ обирається у формі, то instance не потрібен при POST, отримуємо з cleaned_data
        
        form = TechnicalTaskProcessForm(request.POST) # Якщо ТЗ обирається у формі
        # Якщо ТЗ передано через URL:
        # form = TechnicalTaskProcessForm(request.POST, instance=technical_task_instance if technical_task_instance else None)
        # Але оскільки ми оновлюємо конкретні поля, краще отримувати ТЗ з cleaned_data 'technical_task_to_process'

        if form.is_valid():
            task_to_process = form.cleaned_data['technical_task_to_process']
            new_status = form.cleaned_data['new_review_result']
            processed_by_person = form.cleaned_data['processed_by']
            processing_note_text = form.cleaned_data['processing_note']
            # outgoing_number = form.cleaned_data.get('outgoing_number')
            # outgoing_date = form.cleaned_data.get('outgoing_date')

            task_to_process.review_result = new_status
            task_to_process.reviewed_by = processed_by_person
            task_to_process.updated_at = timezone.now()
            # Додаємо примітку про опрацювання до існуючої або створюємо нову
            if processing_note_text:
                if task_to_process.note:
                    task_to_process.note += f"\n--- Опрацювання ({timezone.now().strftime('%d.%m.%Y %H:%M')}) ---\n{processing_note_text}"
                else:
                    task_to_process.note = f"--- Опрацювання ({timezone.now().strftime('%d.%m.%Y %H:%M')}) ---\n{processing_note_text}"
            
            # Якщо є поля для вихідного номера/дати, оновлюємо їх
            # task_to_process.outgoing_document_number = outgoing_number 
            # task_to_process.outgoing_document_date = outgoing_date

            task_to_process.save() # updated_at оновиться автоматично

            messages.success(request, f"Технічне завдання №{task_to_process.input_number} (ОІД: {task_to_process.oid.cipher}) успішно опрацьовано. Новий статус: {task_to_process.get_review_result_display()}.")
            return redirect('oids:list_technical_tasks') 
        else:
            messages.error(request, "Будь ласка, виправте помилки у формі.")
            print(f"TechnicalTaskProcessForm errors: {form.errors.as_json()}")

    else: # GET request
        initial_form_data = {}
        if technical_task_instance:
            initial_form_data['technical_task_to_process'] = technical_task_instance
            # Можна також передати reviewed_by, якщо він є у request.user
            # if request.user.is_authenticated and hasattr(request.user, 'person_profile'):
            #     initial_form_data['processed_by'] = request.user.person_profile

        form = TechnicalTaskProcessForm(initial=initial_form_data)
        # Якщо technical_task_instance визначено, можна обмежити queryset для technical_task_to_process
        # або просто обрати його, якщо поле не disabled
        if technical_task_instance:
            form.fields['technical_task_to_process'].queryset = TechnicalTask.objects.filter(pk=technical_task_instance.pk)
            form.fields['technical_task_to_process'].empty_label = None # Прибираємо порожню опцію, якщо ТЗ вже обрано

    context = {
        'form': form,
        'page_title': f"Опрацювання Технічного Завдання{': ' + str(technical_task_instance) if technical_task_instance else ''}",
        'technical_task_instance': technical_task_instance # Для відображення деталей ТЗ на сторінці
    }
    return render(request, 'oids/forms/technical_task_process_form.html', context)

 
# list info 
@login_required 
def summary_information_hub_view(request):
    """
    Сторінка-хаб з посиланнями на списки об'єктів різних моделей.
    """
    # Список моделей та відповідних їм назв URL для перегляду
    # Ключ 'label' - це те, що побачить користувач
    # Ключ 'url_name' - це name з urls.py для відповідного списку
    model_views = [
        {'label': 'Військові частини', 'url_name': 'oids:list_units'},
        {'label': 'Об\'єкти інформаційної діяльності (ОІД)', 'url_name': 'oids:list_oids'},
        {'label': 'Історія змін статусу ОІД', 'url_name': 'oids:list_oid_status_changes'},
        {'label': 'опрацювання Технічні завдання/Моделі загроз', 'url_name': 'oids:list_technical_tasks'},
        {'label': 'Заявки на проведення робіт', 'url_name': 'oids:list_work_requests'},
        {'label': 'Відрядження', 'url_name': 'oids:list_trips'},
        {'label': 'Надсилання на реєстрацію Актів Атестації', 'url_name': 'oids:list_attestation_registrations'},
        {'label': 'Відповіді Реєстрація Атестації', 'url_name': 'oids:list_attestation_responses'},
        {'label': 'Всі опрацьовані документи', 'url_name': 'oids:list_documents'},
        {'label': 'Всі Акти Атестації (зареєстровані)', 'url_name': 'oids:list_registered_acts'},
        {'label': 'Всі AZR', 'url_name': 'oids:list_azr_documents'},
        {'label': 'Всі Декларації', 'url_name': 'oids:list_declarations'},
        {'label': 'Надсилання до частин пакетів документів', 'url_name': 'oids:list_trip_results_for_units'},
        {'label': 'Довідник: Територіальні управління', 'url_name': 'oids:list_territorial_managements'},
        {'label': 'Довідник: Групи військових частин', 'url_name': 'oids:list_unit_groups'},
        {'label': 'Довідник: Типи документів', 'url_name': 'oids:list_document_types'},
        {'label': 'Довідник: Виконавці (Особи)', 'url_name': 'oids:list_persons'},
    ]
    context = {
        'page_title': 'Зведена інформація та перегляд даних за розділами',
        'model_views': model_views
    }
    return render(request, 'oids/summary_information_hub.html', context)

@login_required 
def document_list_view(request):
	
    documents_list = Document.objects.select_related(
        'oid__unit',
        'document_type', 
        'author',
        'work_request_item__request'
	)
    # ).order_by('-doc_process_date', '-created_at')
	
    form = DocumentFilterForm(request.GET or None)
    if form.is_valid():
        if form.cleaned_data.get('unit'):
            documents_list = documents_list.filter(oid__unit__in=form.cleaned_data['unit'])
        if form.cleaned_data.get('document_type'):
            documents_list = documents_list.filter(document_type__in=form.cleaned_data['document_type'])
        if form.cleaned_data.get('author'):
            documents_list = documents_list.filter(author__in=form.cleaned_data['author'])
        if form.cleaned_data.get('date_from'):
            documents_list = documents_list.filter(doc_process_date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            documents_list = documents_list.filter(doc_process_date__lte=form.cleaned_data['date_to'])

        search_query = form.cleaned_data.get('search_query')
        if search_query:
            documents_list = documents_list.filter(
                Q(document_number__icontains=search_query) |
                Q(oid__cipher__icontains=search_query) |
                Q(note__icontains=search_query)
            ).distinct()
            
	# --- ОНОВЛЕНА ЛОГІКА СОРТУВАННЯ ---
    sort_by = request.GET.get('sort_by', 'doc_process_date') # Ключ для сортування
    sort_order = request.GET.get('sort_order', 'desc')   # Напрямок

    valid_sort_fields = {
        'unit': 'oid__unit__code',
        'oid': 'oid__cipher',
        'doc_type': 'document_type__name',
        'doc_num': 'document_number',
        'proc_date': 'doc_process_date',
        'work_date': 'work_date',
        'exp_date': 'expiration_date',
        'author': 'author__full_name',
    }
    
    order_by_field = valid_sort_fields.get(sort_by)
    
    if order_by_field:
        if sort_order == 'desc':
            order_by_field = f'-{order_by_field}'
        documents_list = documents_list.order_by(order_by_field, '-created_at')
    else:
        # Сортування за замовчуванням
        documents_list = documents_list.order_by('-doc_process_date', '-created_at')
    # --- КІНЕЦЬ БЛОКУ СОРТУВАННЯ ---
    
    # --- ЛОГІКА ЕКСПОРТУ В EXCEL ---
    if request.GET.get('export') == 'excel':
        columns = {
            'oid__unit__code': 'ВЧ',
            'oid__cipher': 'ОІД',
            'document_type__name': 'Тип документа',
            'document_number': 'Підг. №',
            'doc_process_date': 'Підг. від',
            'work_date': 'Дата проведення робіт',
            'expiration_date': 'Термін дії', 
            'author__full_name': 'Автор',
            'note': 'Примітки',
        }
        return export_to_excel(
            documents_list, 
            columns, 
            filename='documents_export.xlsx', 
            include_row_numbers=True
        )

    # --- Пагінація ---
    page_obj = get_paginated_page(documents_list, request)

	# Готуємо параметри для URL-адрес сортування
    query_params = request.GET.copy()
    for key in ['sort_by', 'sort_order', 'page']:
        if key in query_params:
            del query_params[key]

    context = {
        'page_title': 'Список опрацьованих документів',
        'documents': page_obj,
        'page_obj': page_obj,
        'form': form,
        'current_sort_by': sort_by,
        'current_sort_order': sort_order,
        'sort_url_part': query_params.urlencode(),
    }
    return render(request, 'oids/lists/document_list.html', context)

@login_required 
def unit_list_view(request):
    units_list_qs = Unit.objects.select_related(
        'territorial_management'
    ).prefetch_related(
        'unit_groups', 
        'oids' 
    ).annotate(
        oid_count=Count('oids')
    )   
    
    # --- Фільтрація ---
    selected_tm_id_str = request.GET.get('territorial_management')
    current_tm_id_int = None # Буде None або int
    if selected_tm_id_str and selected_tm_id_str.isdigit():
        current_tm_id_int = int(selected_tm_id_str)
        units_list_qs = units_list_qs.filter(territorial_management__id=current_tm_id_int)
    
    search_query = request.GET.get('search_query')
    if search_query:
        units_list_qs = units_list_qs.filter(
            Q(code__icontains=search_query) | 
            Q(name__icontains=search_query) | 
            Q(city__icontains=search_query)
        )

    # --- Сортування ---
    sort_by = request.GET.get('sort_by', 'territorial_management__name')
    sort_order = request.GET.get('sort_order', 'asc')

    valid_sort_fields = {
        'code': 'code',
        'name': 'name',
        'city': 'city',
        'tm': 'territorial_management__name',
        'oid_count': 'oid_count',
        'distance': 'distance_from_gu'
    }
    
    order_by_field = valid_sort_fields.get(sort_by, 'territorial_management__name')

    if sort_order == 'desc':
        order_by_field = f'-{order_by_field}'
    
    units_list_qs = units_list_qs.order_by(order_by_field, 'code' if order_by_field != 'code' else 'name')
    page_obj = get_paginated_page(units_list_qs, request)

    territorial_managements_for_filter = TerritorialManagement.objects.all().order_by('name')

    context = {
        'page_title': 'Список військових частин',
        'units': page_obj,
        'page_obj': page_obj,
        'territorial_managements_for_filter': territorial_managements_for_filter,
        'current_tm_id': current_tm_id_int, # Передаємо int або None
        'current_search_query': search_query,
        'current_sort_by': sort_by,
        'current_sort_order': sort_order,
        'is_sorted_desc': sort_order == 'desc',
    }
    return render(request, 'oids/lists/unit_list.html', context)

@login_required 
def territorial_management_list_view(request):
    tm_list_queryset = TerritorialManagement.objects.all().order_by('name')
    page_obj = get_paginated_page(tm_list_queryset, request)
    
    context = {
        'page_title': 'Список Територіальних Управлінь',
        'object_list': page_obj, # Універсальне ім'я для використання в pagination.html
        'page_obj': page_obj     # Для самого шаблону пагінації
    }
    return render(request, 'oids/lists/territorial_management_list.html', context)

@login_required 
def unit_group_list_view(request):
    group_list_queryset = UnitGroup.objects.prefetch_related('units').order_by('name') # prefetch_related для units
    page_obj = get_paginated_page(group_list_queryset, request)
    
    context = {
        'page_title': 'Список Груп Військових Частин',
        'object_list': page_obj,
        'page_obj': page_obj
    }
    return render(request, 'oids/lists/unit_group_list.html', context)

@login_required 
def person_list_view(request):
    person_list_queryset = Person.objects.all().order_by('full_name')
    page_obj = get_paginated_page(person_list_queryset, request)

        
    context = {
        'page_title': 'Список Виконавців (Осіб)',
        'object_list': page_obj,
        'page_obj': page_obj
    }
    return render(request, 'oids/lists/person_list.html', context)

@login_required 
def document_type_list_view(request):
    doc_type_list_queryset = DocumentType.objects.all().order_by('oid_type', 'work_type', 'name')
    page_obj = get_paginated_page(doc_type_list_queryset, request)
        
    context = {
        'page_title': 'Довідник: Типи документів',
        'object_list': page_obj,
        'page_obj': page_obj
    }
    return render(request, 'oids/lists/document_type_list.html', context)

@login_required 
def oid_list_view(request):
    oid_list_queryset = OID.objects.select_related(
        'unit',  # Завантажуємо пов'язану військову частину
        'unit__territorial_management' # А також ТУ для ВЧ, якщо потрібно (наприклад, для відображення)
    )
	
    # --- Фільтрація ---
    form = OIDFilterForm(request.GET or None)
    if form.is_valid():
        if form.cleaned_data.get('city'):
            oid_list_queryset = oid_list_queryset.filter(unit__city__in=form.cleaned_data['city'])
        if form.cleaned_data.get('unit'):
            oid_list_queryset = oid_list_queryset.filter(unit__in=form.cleaned_data['unit'])
        if form.cleaned_data.get('oid_type'):
            oid_list_queryset = oid_list_queryset.filter(oid_type__in=form.cleaned_data['oid_type'])
        if form.cleaned_data.get('pemin_sub_type'):
            oid_list_queryset = oid_list_queryset.filter(pemin_sub_type__in=form.cleaned_data['pemin_sub_type'])
        if form.cleaned_data.get('status'):
            oid_list_queryset = oid_list_queryset.filter(status__in=form.cleaned_data['status'])
        if form.cleaned_data.get('sec_level'):
            oid_list_queryset = oid_list_queryset.filter(sec_level__in=form.cleaned_data['sec_level'])
        
        search_query = form.cleaned_data.get('search_query')
        if search_query:
            oid_list_queryset = oid_list_queryset.filter(
                Q(cipher__icontains=search_query) |
                Q(full_name__icontains=search_query) |
                Q(room__icontains=search_query) |
                Q(note__icontains=search_query)
            )

    # --- Сортування ---
    # За замовчуванням сортуємо за датою створення (новіші спочатку), якщо поле created_at існує
    # Перевірте вашу модель OID на наявність поля created_at
    # Якщо у вас є поле created_at = models.DateTimeField(auto_now_add=True)
# --- Сортування ---
    sort_by_param = request.GET.get('sort_by') # Прибираємо дефолтне значення тут, тепер воно None
    sort_order_from_request = request.GET.get('sort_order', '')
    
    if sort_by_param:
        # === ВАРІАНТ 1: Користувач обрав конкретне поле для сортування ===
        
        actual_sort_order_is_desc = False
        if sort_by_param.startswith('-'):
            actual_sort_order_is_desc = True
            sort_by_param_cleaned = sort_by_param[1:]
        else:
            sort_by_param_cleaned = sort_by_param
            if sort_order_from_request == 'desc':
                actual_sort_order_is_desc = True
        
        valid_sort_fields = {
            'city': 'unit__city',
            'unit': 'unit__code',
            'cipher': 'cipher',
            'full_name': 'full_name',
            'oid_type': 'oid_type',
            'pemin_sub_type': 'pemin_sub_type',
            'room': 'room',
            'status': 'status',
            'sec_level': 'sec_level',
            'created_at': 'created_at'
        }
        
        # Якщо поле не знайдено, сортуємо за created_at як fallback
        order_by_field_key = valid_sort_fields.get(sort_by_param_cleaned, 'created_at')
        final_order_by_field = f"-{order_by_field_key}" if actual_sort_order_is_desc else order_by_field_key
        
        # Вторинне сортування для стабільності списку
        if order_by_field_key == 'created_at':
            secondary_sort = 'cipher'
        elif order_by_field_key == 'cipher':
            secondary_sort = '-created_at'
        else:
            secondary_sort = '-created_at'

        oid_list_queryset = oid_list_queryset.order_by(final_order_by_field, secondary_sort)
        
    else:
        # === ВАРІАНТ 2: Сортування за замовчуванням (Первинне) ===
        # 1 - за unit (використовуємо unit__code для сортування за номером/назвою, а не ID)
        # 2 - за типом OID
        # 3 - за шифром (cipher)
        oid_list_queryset = oid_list_queryset.order_by('unit__code', 'oid_type', 'cipher')
        
	# --- ЕКСПОРТ В EXCEL ---
    if request.GET.get('export') == 'excel':
        # Стовпці для Excel, що відповідають таблиці на сторінці
        columns = {
            'unit__city': 'Місто',
            'unit__code': 'В/Ч',
            'cipher': 'Шифр ОІД',
            'full_name': 'Повна назва ОІД',
            'get_oid_type_display': 'Тип ОІД',
            'get_pemin_sub_type_display': 'Клас',
            'room': 'Приміщення №',
            'get_status_display': 'Статус',
            'get_sec_level_display': 'Гриф',
            'unit__note': 'Примітка',
        }
        return export_to_excel(
            oid_list_queryset, 
            columns, 
            filename='OID_list_export.xlsx', 
            include_row_numbers=True # Вмикаємо нумерацію
        )
    
    # --- Пагінація ---
    page_obj = get_paginated_page(oid_list_queryset, request)
    
    context = {
        'page_title': 'Список Об\'єктів Інформаційної Діяльності (ОІД)',
        'object_list': page_obj,
        'page_obj': page_obj,
        'form': form, 
        # Додаємо перевірку на None, щоб не виникала помилка при lstrip
        'current_sort_by': sort_by_param.lstrip('-') if sort_by_param else '',
        'current_sort_order_is_desc': sort_by_param.startswith('-') if sort_by_param else False,
    }
    return render(request, 'oids/lists/oid_list.html', context)

@login_required 
def work_request_list_view(request):
    work_request_list_queryset = WorkRequest.objects.select_related(
        'unit', 
        'unit__territorial_management' # Для можливого відображення ТУ
    ).prefetch_related(
        Prefetch('items', queryset=WorkRequestItem.objects.select_related('oid')) 
    )

    # --- ЛОГІКА ФІЛЬТРАЦІЇ ---
    form = WorkRequestFilterForm(request.GET or None)
    if form.is_valid():
        if form.cleaned_data.get('unit'):
            work_request_list_queryset = work_request_list_queryset.filter(unit__in=form.cleaned_data['unit'])
        if form.cleaned_data.get('status'):
            work_request_list_queryset = work_request_list_queryset.filter(status__in=form.cleaned_data['status'])
        if form.cleaned_data.get('date_from'):
            work_request_list_queryset = work_request_list_queryset.filter(incoming_date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            work_request_list_queryset = work_request_list_queryset.filter(incoming_date__lte=form.cleaned_data['date_to'])

        search_query = form.cleaned_data.get('search_query')
        if search_query:
            work_request_list_queryset = work_request_list_queryset.filter(
                Q(incoming_number__icontains=search_query) |
                Q(unit__code__icontains=search_query) |
                Q(unit__name__icontains=search_query) |
                Q(items__oid__cipher__icontains=search_query)
            ).distinct()
            
    # --- Сортування ---
    sort_by_param = request.GET.get('sort_by', '-incoming_date') # За замовчуванням - новіші заявки
    sort_order_from_request = request.GET.get('sort_order', '') 

    actual_sort_order_is_desc = False
    if sort_by_param.startswith('-'):
        actual_sort_order_is_desc = True
        sort_by_param_cleaned = sort_by_param[1:]
    else:
        sort_by_param_cleaned = sort_by_param
        if sort_order_from_request == 'desc':
            actual_sort_order_is_desc = True
            
    valid_sort_fields = {
        'unit': 'unit__code',
        'number': 'incoming_number',
        'date': 'incoming_date',
        'status': 'status',
    }
    
    order_by_field_key = valid_sort_fields.get(sort_by_param_cleaned, 'incoming_date')
    final_order_by_field = f"-{order_by_field_key}" if actual_sort_order_is_desc else order_by_field_key
    
    # Вторинне сортування
    secondary_sort = '-pk' if order_by_field_key != 'incoming_date' else 'unit__code'

    work_request_list_queryset = work_request_list_queryset.order_by(final_order_by_field, secondary_sort)


	# --- КОДУ імпорту excel ---
    if request.GET.get('export') == 'excel':
        # Визначаємо, які стовпці та з якими назвами ми хочемо бачити в Excel
        columns = {
			'unit__code': 'ВЧ',
			'incoming_number': 'Вх. номер заявки',
			'incoming_date': 'Вх. дата заявки',
			'get_status_display': 'Статус заявки',
			'get_items_for_export': 'ОІДи в заявці (Тип робіт / Статус по ОІД)',
            'id': 'ID Заявки',
			'note': 'Примітки',
        }
        # Передаємо ВІДФІЛЬТРОВАНИЙ queryset у нашу функцію
        return export_to_excel(
            work_request_list_queryset, 
            columns, 
            filename='work_requests_export.xlsx', 
            include_row_numbers=True
        )
    # --- КІНЕЦЬ КОДУ імпорту excel ---

    # --- Пагінація ---
    page_obj = get_paginated_page(work_request_list_queryset, request)
        
    context = {
        'page_title': 'Список Заявок на Проведення Робіт',
        'object_list': page_obj,
        'page_obj': page_obj,
        'form': form, # <--- Передаємо екземпляр форми
        'current_sort_by': sort_by_param.lstrip('-'),
        'current_sort_order_is_desc': sort_by_param.startswith('-'),
    }
    return render(request, 'oids/lists/work_request_list.html', context)

@login_required  
def trip_list_view(request):
    trip_list_queryset = Trip.objects.prefetch_related(
        'units', 
        'oids__unit', # Для доступу до oid.unit.code без додаткових запитів
        'persons', 
        'work_requests'
    )

    # --- Фільтрація ---
    filter_unit_id_str = request.GET.get('filter_unit')
    filter_person_id_str = request.GET.get('filter_person')
    filter_date_from_str = request.GET.get('filter_date_from')
    filter_date_to_str = request.GET.get('filter_date_to')
    search_query = request.GET.get('search_query') # Для мети, шифру ОІД, коду ВЧ

    current_filter_unit_id = None
    if filter_unit_id_str and filter_unit_id_str.isdigit():
        current_filter_unit_id = int(filter_unit_id_str)
        # Фільтруємо відрядження, які мають цю ВЧ у своєму списку units
        trip_list_queryset = trip_list_queryset.filter(units__id=current_filter_unit_id) 
    
    current_filter_person_id = None
    if filter_person_id_str and filter_person_id_str.isdigit():
        current_filter_person_id = int(filter_person_id_str)
        # Фільтруємо відрядження, які мають цю особу у своєму списку persons
        trip_list_queryset = trip_list_queryset.filter(persons__id=current_filter_person_id)

    current_filter_date_from = None
    if filter_date_from_str:
        try:
            current_filter_date_from = datetime.strptime(filter_date_from_str, '%Y-%m-%d').date()
            # Фільтр по даті початку АБО даті закінчення (або перетину діапазону)
            trip_list_queryset = trip_list_queryset.filter(
                Q(start_date__gte=current_filter_date_from) | Q(end_date__gte=current_filter_date_from)
            )
        except ValueError:
            current_filter_date_from = None

    current_filter_date_to = None
    if filter_date_to_str:
        try:
            current_filter_date_to = datetime.strptime(filter_date_to_str, '%Y-%m-%d').date()
            trip_list_queryset = trip_list_queryset.filter(
                Q(start_date__lte=current_filter_date_to) | Q(end_date__lte=current_filter_date_to)
            )
        except ValueError:
            current_filter_date_to = None
            
    if search_query:
        trip_list_queryset = trip_list_queryset.filter(
            Q(purpose__icontains=search_query) |
            Q(units__code__icontains=search_query) |
            Q(oids__cipher__icontains=search_query) |
            Q(persons__full_name__icontains=search_query) |
            Q(work_requests__incoming_number__icontains=search_query)
        ).distinct() # distinct важливий через пошук по M2M

    # --- Сортування ---
    sort_by_param = request.GET.get('sort_by', '-start_date') # За замовчуванням - новіші відрядження
    sort_order_from_request = request.GET.get('sort_order', '') 

    actual_sort_order_is_desc = False
    if sort_by_param.startswith('-'):
        actual_sort_order_is_desc = True
        sort_by_param_cleaned = sort_by_param[1:]
    else:
        sort_by_param_cleaned = sort_by_param
        if sort_order_from_request == 'desc':
            actual_sort_order_is_desc = True
            
    valid_sort_fields = {
        'start_date': 'start_date',
        'end_date': 'end_date',
        'purpose': 'purpose',
        # Сортування за M2M полями (units, oids, persons) напряму через order_by складне.
        # Якщо потрібно, це вимагає анотації або більш складних запитів.
        # Поки що обмежимось прямими полями моделі Trip.
    }
    
    order_by_field_key = valid_sort_fields.get(sort_by_param_cleaned, 'start_date')
    final_order_by_field = f"-{order_by_field_key}" if actual_sort_order_is_desc else order_by_field_key
    
    secondary_sort = '-pk' # Загальне вторинне сортування для стабільності

    trip_list_queryset = trip_list_queryset.order_by(final_order_by_field, secondary_sort).distinct() # distinct тут теж може бути корисним

    # --- Пагінація ---
    page_obj = get_paginated_page(trip_list_queryset, request)
        
    # Дані для фільтрів
    units_for_filter = Unit.objects.all().order_by('code')
    persons_for_filter = Person.objects.filter(is_active=True).order_by('full_name')
        
    context = {
        'page_title': 'Список Відряджень',
        'object_list': page_obj,
        'page_obj': page_obj,
        # Фільтри
        'units_for_filter': units_for_filter,
        'persons_for_filter': persons_for_filter,
        'current_filter_unit_id': current_filter_unit_id,
        'current_filter_person_id': current_filter_person_id,
        'current_filter_date_from': filter_date_from_str,
        'current_filter_date_to': filter_date_to_str,
        'current_search_query': search_query,
        # Сортування
        'current_sort_by': sort_by_param_cleaned,
        'current_sort_order_is_desc': actual_sort_order_is_desc,
    }
    return render(request, 'oids/lists/trip_list.html', context)

@login_required 
def technical_task_list_view(request):
    task_list_queryset = TechnicalTask.objects.select_related(
        'oid__unit', 
        'oid',
        'reviewed_by'
    )

    # --- Фільтрація ---
    form = TechnicalTaskFilterForm(request.GET or None)
    if form.is_valid():
        if form.cleaned_data.get('unit'):
            task_list_queryset = task_list_queryset.filter(oid__unit__in=form.cleaned_data['unit'])
        if form.cleaned_data.get('oid'):
            task_list_queryset = task_list_queryset.filter(oid__in=form.cleaned_data['oid'])
        if form.cleaned_data.get('review_result'):
            task_list_queryset = task_list_queryset.filter(review_result__in=form.cleaned_data['review_result'])
        if form.cleaned_data.get('reviewed_by'):
            task_list_queryset = task_list_queryset.filter(reviewed_by__in=form.cleaned_data['reviewed_by'])
        if form.cleaned_data.get('input_date_from'):
            task_list_queryset = task_list_queryset.filter(input_date__gte=form.cleaned_data['input_date_from'])
        if form.cleaned_data.get('input_date_to'):
            task_list_queryset = task_list_queryset.filter(input_date__lte=form.cleaned_data['input_date_to'])
        if form.cleaned_data.get('read_till_date_from'):
            task_list_queryset = task_list_queryset.filter(read_till_date__gte=form.cleaned_data['read_till_date_from'])
        if form.cleaned_data.get('read_till_date_to'):
            task_list_queryset = task_list_queryset.filter(read_till_date__lte=form.cleaned_data['read_till_date_to'])

        search_query = form.cleaned_data.get('search_query')
        if search_query:
            task_list_queryset = task_list_queryset.filter(
                Q(input_number__icontains=search_query) |
                Q(oid__cipher__icontains=search_query) |
                Q(oid__full_name__icontains=search_query) |
                Q(note__icontains=search_query)
            ).distinct()

    # --- Сортування ---
    sort_by_param = request.GET.get('sort_by', '-input_date') # За замовчуванням, як у вас було
    sort_order_from_request = request.GET.get('sort_order', '') 

    actual_sort_order_is_desc = False
    if sort_by_param.startswith('-'):
        actual_sort_order_is_desc = True
        sort_by_param_cleaned = sort_by_param[1:]
    else:
        sort_by_param_cleaned = sort_by_param
        if sort_order_from_request == 'desc':
            actual_sort_order_is_desc = True
            
    valid_sort_fields = {
        'oid_loc': 'oid__unit__code', # Сортування за кодом ВЧ, потім можна додати шифр ОІД
        'input_number': 'input_number', #
        'input_date': 'input_date', #
        'read_till_date': 'read_till_date', #
        'review_result': 'review_result', #
		'updated_at': 'updated_at',
        'reviewed_by': 'reviewed_by__full_name',
        'created_at': 'created_at',
    }
    
    order_by_field_key = valid_sort_fields.get(sort_by_param_cleaned, 'input_date')
    final_order_by_field = f"-{order_by_field_key}" if actual_sort_order_is_desc else order_by_field_key
    
    # Вторинне сортування
    if order_by_field_key == 'input_date':
        secondary_sort = '-created_at' #
    elif order_by_field_key == 'oid__unit__code':
        secondary_sort = 'oid__cipher' # Якщо сортуємо за ВЧ, то потім за шифром ОІД
    else:
        secondary_sort = '-input_date'

    task_list_queryset = task_list_queryset.order_by(final_order_by_field, secondary_sort).distinct()

    # --- ЕКСПОРТ В EXCEL ---
    if request.GET.get('export') == 'excel':
        columns = {
            'oid__unit__code': 'ВЧ', 
            'oid__cipher': 'ОІД (Шифр)',
            'input_number': 'Вхідний № ТЗ',
            'input_date': 'Вхідна дата',
            'read_till_date': 'Опрацювати ДО',
            'get_review_result_display': 'Статус',
            'reviewed_by__full_name': 'Хто опрацював',
            'updated_at': 'Дата опрацювання',
            'note': 'Примітки',
        }
        return export_to_excel(
            task_list_queryset, 
            columns, 
            filename='technical_tasks_export.xlsx', 
            include_row_numbers=True
        )

    # --- Пагінація ---
    page_obj = get_paginated_page(task_list_queryset, request)

    context = {
        'page_title': 'Список опрацювання ТЗ/МЗ',
        'object_list': page_obj,
        'page_obj': page_obj,
        'form': form, # Передаємо форму в шаблон
        'current_sort_by': sort_by_param.lstrip('-'),
        'current_sort_order_is_desc': sort_by_param.startswith('-'),
    }
    return render(request, 'oids/lists/technical_task_list.html', context)


# 1. View для СТВОРЕННЯ відправки у в/ч (основний)
@login_required
def send_trip_results_form(request):
    """
    Форма для створення відправки результатів відрядження у в/ч
    """
    if request.method == 'POST':
        form = TripResultForUnitForm(request.POST)
        
        if form.is_valid():
            trip_result = form.save(commit=False)
            # Можна встановити хто створив/відправив
            # trip_result.created_by = request.user.person
            trip_result.save()
            
            # Зберігаємо M2M зв'язки
            form.save_m2m()
            
            # ⭐ КЛЮЧОВИЙ МОМЕНТ: Оновлюємо статуси після відправки
            update_statuses_after_sending_to_unit(trip_result)
            
            messages.success(
                request, 
                f'✅ Результати відрядження успішно відправлено у в/ч. '
                f'Оновлено статуси заявок.'
            )
            return redirect('oids:trip_result_for_unit_list')
    else:
        form = TripResultForUnitForm()
    
    context = {
        'form': form,
        'page_title': 'Відправити документи в частину',
    }
    return render(request, 'oids/forms/send_trip_results_form.html', context)


# 2. НОВА ФУНКЦІЯ для оновлення статусів
def update_statuses_after_sending_to_unit(trip_result):
    """
    Оновлює статуси WorkRequestItem після відправки документів у в/ч
    
    Args:
        trip_result: Об'єкт TripResultForUnit
    """
    print(f"\n{'='*60}")
    print(f"[TRIP_RESULT_STATUS_UPDATE] Processing TripResultForUnit ID {trip_result.id}")
    print(f"Outgoing letter: №{trip_result.outgoing_letter_number} від {trip_result.outgoing_letter_date}")
    print(f"{'='*60}\n")
    
    updated_wri_count = 0
    documents_processed = 0
    
    # Обробляємо всі документи у відправці
    for document in trip_result.documents.all():
        documents_processed += 1
        print(f"\n📄 Processing Document ID {document.id}")
        print(f"   Type: {document.document_type.name if document.document_type else 'Unknown'}")
        print(f"   OID: {document.oid.cipher}")
        print(f"   Number: {document.document_number}")
        
        # Перевіряємо чи є пов'язаний WorkRequestItem
        if document.work_request_item:
            wri = document.work_request_item
            print(f"   ✅ Linked to WRI ID {wri.id} (Status: {wri.get_status_display()})")
            
            # Викликаємо оновлення статусу
            old_status = wri.status
            wri.check_and_update_status_based_on_documents()
            
            if wri.status != old_status:
                updated_wri_count += 1
                print(f"   🔄 Status changed: {old_status} → {wri.status}")
            else:
                print(f"   ℹ️  Status unchanged: {wri.status}")
        else:
            print(f"   ⚠️  No linked WorkRequestItem")
    
    print(f"\n{'='*60}")
    print(f"[TRIP_RESULT_STATUS_UPDATE] Summary:")
    print(f"  Documents processed: {documents_processed}")
    print(f"  WorkRequestItems updated: {updated_wri_count}")
    print(f"{'='*60}\n")


# 3. View для списку (без змін, але з коментарями)
@login_required 
def trip_result_for_unit_list_view(request):
    """
    Список відправок результатів відряджень у в/ч
    """
    result_list_queryset = TripResultForUnit.objects.select_related(
        'trip'
    ).prefetch_related(
        'units',
        Prefetch('oids', queryset=OID.objects.select_related('unit')),
        Prefetch('documents', queryset=Document.objects.select_related(
            'document_type',
            'work_request_item',  # ДОДАНО для перевірки зв'язку
            'work_request_item__request'  # ДОДАНО для перевірки заявки
        ))
    ).order_by('-outgoing_letter_date')

    page_obj = get_paginated_page(result_list_queryset, request)
        
    context = {
        'page_title': 'Список Результатів Відряджень (відправка до ВЧ)',
        'object_list': page_obj,
        'page_obj': page_obj
    }
    return render(request, 'oids/lists/trip_result_for_unit_list.html', context)




# 7. TESTING: Функція для тестування
def test_trip_result_status_update(trip_result_id):
    """
    Тестова функція для перевірки оновлення статусів
    Використання: test_trip_result_status_update(1)
    """
    from oids.models import TripResultForUnit
    
    trip_result = TripResultForUnit.objects.get(id=trip_result_id)
    
    print("\n" + "="*60)
    print(f"TESTING: TripResultForUnit ID {trip_result_id}")
    print("="*60)
    
    for doc in trip_result.documents.all():
        print(f"\nDocument: {doc.document_number}")
        print(f"  Type: {doc.document_type.name if doc.document_type else 'N/A'}")
        print(f"  OID: {doc.oid.cipher}")
        
        if doc.work_request_item:
            wri = doc.work_request_item
            print(f"  WRI ID: {wri.id}")
            print(f"  Current Status: {wri.get_status_display()}")
            print(f"  Work Type: {wri.get_work_type_display()}")
            
            # Перевірка умов
            print(f"\n  Checking conditions:")
            print(f"    - doc_process_date: {doc.doc_process_date}")
            print(f"    - is_sent_to_unit: {doc.is_sent_to_unit}")
            print(f"    - trip_result_sent: {doc.trip_result_sent}")
            
            if wri.work_type in ['ATTESTATION', 'PLAND_ATTESTATION']:
                print(f"    - dsszzi_registered_number: {doc.dsszzi_registered_number}")
                print(f"    - attestation_registration_sent: {doc.attestation_registration_sent}")
            
            # Викликаємо оновлення
            print(f"\n  Calling check_and_update_status_based_on_documents()...")
            wri.check_and_update_status_based_on_documents()
            wri.refresh_from_db()
            print(f"  New Status: {wri.get_status_display()}")
        else:
            print(f"  ⚠️  No WorkRequestItem linked")
    
    print("\n" + "="*60 + "\n")




# У відповідному view після збереження TripResultForUnit
def trip_result_send_view(request):
    # ... existing code ...
    
    if form.is_valid():
        trip_result = form.save()
        
        # Встановлюємо зв'язок trip_result_sent для кожного документа
        for doc in trip_result.documents.all():
            if not doc.trip_result_sent:
                doc.trip_result_sent = trip_result
                doc.save(update_fields=['trip_result_sent', 'updated_at'])
        
        # Перевіряємо статуси WorkRequestItems
        for doc in trip_result.documents.all():
            if doc.work_request_item:
                doc.work_request_item.check_and_update_status_based_on_documents()                
				
@login_required 
def oid_status_change_list_view(request):
    status_change_list_queryset = OIDStatusChange.objects.select_related(
        'oid__unit', 
        'oid',
        'initiating_document__document_type', 
        'changed_by' 
    )

    # --- Фільтрація ---
    filter_unit_id_str = request.GET.get('filter_unit')
    filter_oid_id_str = request.GET.get('filter_oid')
    filter_old_status = request.GET.get('filter_old_status')
    filter_new_status = request.GET.get('filter_new_status')
    filter_changed_by_id_str = request.GET.get('filter_changed_by')
    filter_date_from_str = request.GET.get('filter_date_from')
    filter_date_to_str = request.GET.get('filter_date_to')
    search_query = request.GET.get('search_query') # Для ОІД, причини, документа

    current_filter_unit_id = None
    if filter_unit_id_str and filter_unit_id_str.isdigit():
        current_filter_unit_id = int(filter_unit_id_str)
        status_change_list_queryset = status_change_list_queryset.filter(oid__unit__id=current_filter_unit_id)

    current_filter_oid_id = None
    if filter_oid_id_str and filter_oid_id_str.isdigit():
        current_filter_oid_id = int(filter_oid_id_str)
        status_change_list_queryset = status_change_list_queryset.filter(oid__id=current_filter_oid_id)
    
    if filter_old_status:
        status_change_list_queryset = status_change_list_queryset.filter(old_status=filter_old_status)
    
    if filter_new_status:
        status_change_list_queryset = status_change_list_queryset.filter(new_status=filter_new_status)

    current_filter_changed_by_id = None
    if filter_changed_by_id_str and filter_changed_by_id_str.isdigit():
        current_filter_changed_by_id = int(filter_changed_by_id_str)
        status_change_list_queryset = status_change_list_queryset.filter(changed_by__id=current_filter_changed_by_id)

    current_filter_date_from = None
    if filter_date_from_str:
        try:
            current_filter_date_from = datetime.strptime(filter_date_from_str, '%Y-%m-%d').date()
            status_change_list_queryset = status_change_list_queryset.filter(changed_at__date__gte=current_filter_date_from)
        except ValueError:
            current_filter_date_from = None

    current_filter_date_to = None
    if filter_date_to_str:
        try:
            current_filter_date_to = datetime.strptime(filter_date_to_str, '%Y-%m-%d').date()
            status_change_list_queryset = status_change_list_queryset.filter(changed_at__date__lte=current_filter_date_to)
        except ValueError:
            current_filter_date_to = None
            
    if search_query:
        status_change_list_queryset = status_change_list_queryset.filter(
            Q(oid__cipher__icontains=search_query) |
            Q(oid__unit__code__icontains=search_query) |
            Q(reason__icontains=search_query) |
            Q(initiating_document__document_number__icontains=search_query) |
            Q(initiating_document__document_type__name__icontains=search_query)
        ).distinct()

    # --- Сортування ---
    sort_by_param = request.GET.get('sort_by', '-changed_at') # За замовчуванням
    sort_order_from_request = request.GET.get('sort_order', '') 

    actual_sort_order_is_desc = False
    if sort_by_param.startswith('-'):
        actual_sort_order_is_desc = True
        sort_by_param_cleaned = sort_by_param[1:]
    else:
        sort_by_param_cleaned = sort_by_param
        if sort_order_from_request == 'desc':
            actual_sort_order_is_desc = True
            
    valid_sort_fields = {
        'oid_loc': 'oid__unit__code', 
        'changed_at': 'changed_at',
        'old_status': 'old_status',
        'new_status': 'new_status',
        'reason': 'reason', # Сортування за причиною може бути менш корисним
        'document': 'initiating_document__document_number', # Або initiating_document__document_type__name
        'changed_by': 'changed_by__full_name',
    }
    
    order_by_field_key = valid_sort_fields.get(sort_by_param_cleaned, 'changed_at')
    final_order_by_field = f"-{order_by_field_key}" if actual_sort_order_is_desc else order_by_field_key
    
    secondary_sort = '-pk' if order_by_field_key != 'changed_at' else 'oid__cipher'

    status_change_list_queryset = status_change_list_queryset.order_by(final_order_by_field, secondary_sort).distinct()

    # --- Пагінація ---
    page_obj = get_paginated_page(status_change_list_queryset, request)
        
    # Дані для фільтрів
    units_for_filter = Unit.objects.all().order_by('code')
    oids_for_filter = OID.objects.select_related('unit').all().order_by('unit__code', 'cipher')
    # Припускаємо, що old_status та new_status використовують ті ж choices, що й статус ОІД
    status_choices_for_filter = OIDStatusChoices.choices 
    persons_for_filter = Person.objects.filter(oid_status_changes__isnull=False).distinct().order_by('full_name')
        
    context = {
        'page_title': 'Історія Змін Статусу ОІД',
        'object_list': page_obj,
        'page_obj': page_obj,
        # Фільтри
        'units_for_filter': units_for_filter,
        'oids_for_filter': oids_for_filter,
        'status_choices_for_filter': status_choices_for_filter, # Однаковий для old_status та new_status
        'persons_for_filter': persons_for_filter,
        'current_filter_unit_id': current_filter_unit_id,
        'current_filter_oid_id': current_filter_oid_id,
        'current_filter_old_status': filter_old_status,
        'current_filter_new_status': filter_new_status,
        'current_filter_changed_by_id': current_filter_changed_by_id,
        'current_filter_date_from': filter_date_from_str,
        'current_filter_date_to': filter_date_to_str,
        'current_search_query': search_query,
        # Сортування
        'current_sort_by': sort_by_param_cleaned,
        'current_sort_order_is_desc': actual_sort_order_is_desc,
    }
    return render(request, 'oids/lists/oid_status_change_list.html', context)

# oids/views.py

@login_required
def attestation_registration_list_view(request):
    """
    Список відправок на реєстрацію атестаційних актів
    """
    queryset = AttestationRegistration.objects.select_related(
        'sent_by'
    ).prefetch_related(
        'units',
        'sent_documents',  # ВИПРАВЛЕНО: було 'registered_documents'
        'sent_documents__oid',
        'sent_documents__oid__unit',
        'sent_documents__document_type'
    ).order_by('-outgoing_letter_date', '-created_at')
    
    # Фільтрація
    filter_form = AttestationRegistrationFilterForm(request.GET or None)
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('units'):  # ВИПРАВЛЕНО: було 'unit'
            queryset = queryset.filter(units__in=filter_form.cleaned_data['units']).distinct()
        if filter_form.cleaned_data.get('status'):
            queryset = queryset.filter(status=filter_form.cleaned_data['status'])
        if filter_form.cleaned_data.get('sent_by'):
            queryset = queryset.filter(sent_by=filter_form.cleaned_data['sent_by'])
        if filter_form.cleaned_data.get('date_from'):
            queryset = queryset.filter(outgoing_letter_date__gte=filter_form.cleaned_data['date_from'])
        if filter_form.cleaned_data.get('date_to'):
            queryset = queryset.filter(outgoing_letter_date__lte=filter_form.cleaned_data['date_to'])
        if filter_form.cleaned_data.get('search_query'):
            search = filter_form.cleaned_data['search_query']
            queryset = queryset.filter(
                Q(outgoing_letter_number__icontains=search) |
                Q(sent_documents__oid__cipher__icontains=search)
            ).distinct()
    
    # Експорт в Excel
    if request.GET.get('export') == 'excel':
        return export_attestation_registrations_to_excel(queryset)
    
    # Пагінація
    page_obj = get_paginated_page(queryset, request)
    
    context = {
        'page_title': 'Реєстрація атестаційних актів: Список відправок',
        'object_list': page_obj,  # ВИПРАВЛЕНО: було 'attestation_registrations'
        'form': filter_form,  # ВИПРАВЛЕНО: було 'filter_form'
        'page_obj': page_obj,  # Для пагінації
    }
    
    return render(request, 'oids/lists/attestation_registration_list.html', context)


# Додаткова функція для експорту (якщо потрібна)
def export_attestation_registrations_to_excel(queryset):
    """Експорт відправок в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Реєстрація АА"
    
    # Заголовки
    headers = [
        'Вихідний №',
        'Дата відправки',
        'ВЧ у відправці',
        'Акти на реєстрацію',
        'Хто відправив',
        'Статус',
        'Типи документів'
    ]
    
    # Стилі для заголовків
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Дані
    for row_num, reg in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = reg.outgoing_letter_number
        ws.cell(row=row_num, column=2).value = reg.outgoing_letter_date.strftime('%d.%m.%Y') if reg.outgoing_letter_date else '-'
        
        # ВЧ у відправці
        units = ', '.join([unit.code for unit in reg.units.all()])
        ws.cell(row=row_num, column=3).value = units or '-'
        
        # Акти на реєстрацію
        acts = ', '.join([f"{doc.oid.cipher} (підг. № {doc.document_number})" for doc in reg.sent_documents.all()])
        ws.cell(row=row_num, column=4).value = acts or '-'
        
        # Хто відправив
        ws.cell(row=row_num, column=5).value = reg.sent_by.full_name if reg.sent_by else '-'
        
        # Статус
        ws.cell(row=row_num, column=6).value = reg.get_status_display()
        
        # Типи документів
        doc_types = ', '.join(set([doc.document_type.name for doc in reg.sent_documents.all() if doc.document_type]))
        ws.cell(row=row_num, column=7).value = doc_types or '-'
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Відправка файлу
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attestation_registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

# @login_required 
# def attestation_registration_list_view(request):
#     registration_list_queryset = AttestationRegistration.objects.select_related(
#         'sent_by'
#     ).prefetch_related(
#         'units', 
#         # Одразу завантажуємо пов'язані документи та їхні ОІДи, щоб уникнути N+1 запитів
#         Prefetch('registered_documents', queryset=Document.objects.select_related('oid__unit'))
#     ).order_by('-outgoing_letter_date', '-id')

#     # --- НОВА ЛОГІКА ФІЛЬТРАЦІЇ ---
#     form = AttestationRegistrationFilterForm(request.GET or None)
#     if form.is_valid():
#         if form.cleaned_data.get('units'):
#             registration_list_queryset = registration_list_queryset.filter(units__in=form.cleaned_data['units']).distinct()
#         if form.cleaned_data.get('status'):
#             registration_list_queryset = registration_list_queryset.filter(status__in=form.cleaned_data['status'])
#         if form.cleaned_data.get('sent_by'):
#             registration_list_queryset = registration_list_queryset.filter(sent_by__in=form.cleaned_data['sent_by'])
#         if form.cleaned_data.get('date_from'):
#             registration_list_queryset = registration_list_queryset.filter(outgoing_letter_date__gte=form.cleaned_data['date_from'])
#         if form.cleaned_data.get('date_to'):
#             registration_list_queryset = registration_list_queryset.filter(outgoing_letter_date__lte=form.cleaned_data['date_to'])

#         search_query = form.cleaned_data.get('search_query')
#         if search_query:
#             registration_list_queryset = registration_list_queryset.filter(
#                 Q(outgoing_letter_number__icontains=search_query) |
#                 Q(note__icontains=search_query)
#             )

# 	# --- ДОДАЄМО НОВУ ЛОГІКУ ЕКСПОРТУ ---
#     if request.GET.get('export') == 'excel':
#         columns = {
#             'outgoing_letter_number': 'Вихідний №',
#             'outgoing_letter_date': 'Дата вих. листа',
#             'get_units_for_export': 'ВЧ у відправці',
#             'get_documents_for_export': 'Акти, надіслані на реєстр.',
#             'sent_by__full_name': 'Хто відправив',
#             'get_status_display': 'Статус',
#         }
#         return export_to_excel(
#             registration_list_queryset, 
#             columns, 
#             filename='attestation_registrations_export.xlsx',
#             include_row_numbers=True
#         )
#     # --- КІНЕЦЬ БЛОКУ ЕКСПОРТУ ---
#     page_obj = get_paginated_page(registration_list_queryset, request)

#     context = {
#         'page_title': 'Відправки Актів Атестації на реєстрацію (ДССЗЗІ)',
#         'object_list': page_obj,
#         'page_obj': page_obj,
#         'form': form, # Передаємо форму в шаблон
#     }
#     return render(request, 'oids/lists/attestation_registration_list.html', context)


# --- ДОДАТКОВА ФУНКЦІЯ ДЛЯ ВІДЛАГОДЖЕННЯ ---
def debug_document_status(work_request_item):
    """
    Функція для відлагодження статусу документів
    """
    print("\n" + "="*60)
    print(f"DEBUG: WorkRequestItem ID {work_request_item.id}")
    print(f"OID: {work_request_item.oid.cipher}")
    print(f"Work Type: {work_request_item.get_work_type_display()}")
    print(f"Current Status: {work_request_item.get_status_display()}")
    print("="*60)
    
    docs = Document.objects.filter(work_request_item=work_request_item)
    
    if not docs.exists():
        print("❌ No documents found")
        return
    
    for doc in docs:
        print(f"\n📄 Document: {doc.document_type.name if doc.document_type else 'Unknown'}")
        print(f"   - Processed Date: {doc.doc_process_date or '❌'}")
        print(f"   - Registration Sent: {'✅' if hasattr(doc, 'attestation_registration_sent') and doc.attestation_registration_sent else '❌'}")
        print(f"   - DSSZZI Number: {doc.dsszzi_registered_number or '❌'}")
        print(f"   - DSSZZI Date: {doc.dsszzi_registered_date or '❌'}")
        print(f"   - Sent to Unit: {'✅' if hasattr(doc, 'trip_result_sent') and doc.trip_result_sent else '❌'}")
    
    print("\n" + "="*60 + "\n")


@login_required
def attestation_response_list_view(request):
    """
    Список відповідей на відправки атестаційних актів
    """
    # Отримуємо фільтр з GET-параметрів
    current_filter_att_reg_id = request.GET.get('attestation_registration_sent', None)
    
    queryset = AttestationResponse.objects.select_related(
        'attestation_registration_sent',
        'attestation_registration_sent__sent_by',
        'received_by'
    ).prefetch_related(
        'attestation_registration_sent__units',
        'attestation_registration_sent__sent_documents',  # ВИПРАВЛЕНО: було 'registered_documents'
        'attestation_registration_sent__sent_documents__oid',
        'attestation_registration_sent__sent_documents__oid__unit',
        'attestation_registration_sent__sent_documents__document_type'
    ).order_by('-response_letter_date', '-created_at')
    
    # Фільтрація
    filter_form = AttestationResponseFilterForm(request.GET or None)
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('attestation_registration_sent'):
            queryset = queryset.filter(
                attestation_registration_sent=filter_form.cleaned_data['attestation_registration_sent']
            )
        if filter_form.cleaned_data.get('received_by'):
            queryset = queryset.filter(received_by=filter_form.cleaned_data['received_by'])
        if filter_form.cleaned_data.get('date_from'):
            queryset = queryset.filter(response_letter_date__gte=filter_form.cleaned_data['date_from'])
        if filter_form.cleaned_data.get('date_to'):
            queryset = queryset.filter(response_letter_date__lte=filter_form.cleaned_data['date_to'])
        if filter_form.cleaned_data.get('search_query'):
            search = filter_form.cleaned_data['search_query']
            queryset = queryset.filter(
                Q(response_letter_number__icontains=search) |
                Q(attestation_registration_sent__outgoing_letter_number__icontains=search) |
                Q(attestation_registration_sent__sent_documents__oid__cipher__icontains=search)
            ).distinct()
    
    # Експорт в Excel
    if request.GET.get('export') == 'excel':
        return export_attestation_responses_to_excel(queryset)
    
    # Пагінація
    page_obj = get_paginated_page(queryset, request)
    
    context = {
        'page_title': 'Реєстрація атестаційних актів: Список відповідей',
        'object_list': page_obj,  # ВИПРАВЛЕНО: було 'attestation_responses'
        'form': filter_form,  # ВИПРАВЛЕНО: було 'filter_form'
        'page_obj': page_obj,  # Для пагінації
        'current_filter_att_reg_id': current_filter_att_reg_id,
    }
    
    return render(request, 'oids/lists/attestation_response_list.html', context)


# Функція експорту в Excel (бонус)
def export_attestation_responses_to_excel(queryset):
    """Експорт відповідей в Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Відповіді ДССЗЗІ"
    
    # Заголовки
    headers = [
        '№',
        'Вхідний лист №',
        'Дата відповіді',
        'На вихідний лист №',
        'Дата вихідного',
        'Акти атестації',
        'Реєстраційні номери ДССЗЗІ',
        'Хто отримав/вніс',
        'Дата внесення'
    ]
    
    # Стилі для заголовків
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Дані
    for row_num, response in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = response.response_letter_number
        ws.cell(row=row_num, column=3).value = response.response_letter_date.strftime('%d.%m.%Y') if response.response_letter_date else '-'
        
        # Вихідний лист
        if response.attestation_registration_sent:
            ws.cell(row=row_num, column=4).value = response.attestation_registration_sent.outgoing_letter_number
            ws.cell(row=row_num, column=5).value = response.attestation_registration_sent.outgoing_letter_date.strftime('%d.%m.%Y') if response.attestation_registration_sent.outgoing_letter_date else '-'
            
            # Акти атестації
            acts = []
            registrations = []
            for doc in response.attestation_registration_sent.sent_documents.all():
                acts.append(f"{doc.oid.cipher} - Акт №{doc.document_number}")
                if doc.dsszzi_registered_number:
                    registrations.append(f"№{doc.dsszzi_registered_number} від {doc.dsszzi_registered_date.strftime('%d.%m.%Y') if doc.dsszzi_registered_date else '-'}")
                else:
                    registrations.append("Не зареєстровано")
            
            ws.cell(row=row_num, column=6).value = '; '.join(acts) if acts else '-'
            ws.cell(row=row_num, column=7).value = '; '.join(registrations) if registrations else '-'
        else:
            ws.cell(row=row_num, column=4).value = '-'
            ws.cell(row=row_num, column=5).value = '-'
            ws.cell(row=row_num, column=6).value = '-'
            ws.cell(row=row_num, column=7).value = '-'
        
        # Хто отримав
        ws.cell(row=row_num, column=8).value = response.received_by.full_name if response.received_by else '-'
        
        # Дата внесення
        ws.cell(row=row_num, column=9).value = response.created_at.strftime('%d.%m.%Y %H:%M') if response.created_at else '-'
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Відправка файлу
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attestation_responses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def attestation_registered_acts_list_view(request):
    # 1. Початковий запит: обираємо тільки Акти Атестації, які зареєстровані
    try:
	# 	        attestation_act_type = DocumentType.objects.get(
    #     duration_months=60,
    #     # Додайте інші поля для унікальності, наприклад:
    #     name="Акт атестації комплексу ТЗІ"
    # )
    # except DocumentType.DoesNotExist:
    #         attestation_act_type = DocumentType.objects.filter(duration_months=60).first() # Якщо все ще повертається кілька об'єктів
    #         pass
    # except DocumentType.MultipleObjectsReturned: 
    #     attestation_act_type = DocumentType.objects.filter(duration_months=60).first() # Якщо все ще повертається кілька об'єктів
        attestation_act_type = DocumentType.objects.get(name__icontains="Акт атестації")
        base_queryset = Document.objects.filter(
            document_type=attestation_act_type,
            dsszzi_registered_number__isnull=False
        ).exclude(dsszzi_registered_number__exact='').select_related(
            'oid__unit', 
            'attestation_registration_sent', 
            'attestation_registration_sent__response_received'
        )
    except DocumentType.DoesNotExist:
        base_queryset = Document.objects.none()

    # 2. Фільтрація
    form = RegisteredActsFilterForm(request.GET or None)
    if form.is_valid():
        if form.cleaned_data.get('unit'):
            base_queryset = base_queryset.filter(oid__unit__in=form.cleaned_data['unit'])
        if form.cleaned_data.get('registration_date_from'):
            base_queryset = base_queryset.filter(dsszzi_registered_date__gte=form.cleaned_data['registration_date_from'])
        if form.cleaned_data.get('registration_date_to'):
            base_queryset = base_queryset.filter(dsszzi_registered_date__lte=form.cleaned_data['registration_date_to'])

        search_query = form.cleaned_data.get('search_query')
        if search_query:
            base_queryset = base_queryset.filter(
                Q(document_number__icontains=search_query) |
                Q(oid__cipher__icontains=search_query) |
                Q(dsszzi_registered_number__icontains=search_query)
            )

    # 3. Сортування
    # ---  ПОВНОЦІННА ЛОГІКА СОРТУВАННЯ ---
    sort_by = request.GET.get('sort_by', 'reg_date') # Ключ для сортування
    sort_order = request.GET.get('sort_order', 'desc') # Напрямок

    valid_sort_fields = {
        'unit': 'oid__unit__code',
        'oid': 'oid__cipher',
        'prep_num': 'document_number',
        'work_date': 'work_date',
        'reg_num': 'dsszzi_registered_number',
        'reg_date': 'dsszzi_registered_date',
    }
    
    order_by_field = valid_sort_fields.get(sort_by)

    if order_by_field:
        if sort_order == 'desc':
            order_by_field = f'-{order_by_field}'
        final_queryset = base_queryset.order_by(order_by_field)
    else:
        # Сортування за замовчуванням, якщо передано невалідний параметр
        final_queryset = base_queryset.order_by('-dsszzi_registered_date')
    # --- КІНЕЦЬ БЛОКУ СОРТУВАННЯ ---

    # 4. Експорт в Excel
    if request.GET.get('export') == 'excel':
        columns = {
            'oid__unit__code': 'ВЧ', 
            'oid__cipher': 'ОІД (Шифр)',
            'document_number': 'Акт атестації (підг. №)',
            'work_date': 'Дата підг. акту',
            'get_sent_info_for_export': 'Відправлено (Вих. лист)',
            'dsszzi_registered_number': 'Номер реєстрації ДССЗЗІ',
            'dsszzi_registered_date': 'Дата реєстрації ДССЗЗІ',
            'get_response_info_for_export': 'Лист-відповідь ДССЗЗІ',
        }
        return export_to_excel(
            final_queryset, 
            columns, 
            filename='registered_acts.xlsx', 
            include_row_numbers=True
        )

    # 5. Пагінація
    page_obj = get_paginated_page(final_queryset, request)

	# Готуємо параметри для URL-адрес сортування
    query_params = request.GET.copy()
    for key in ['sort_by', 'sort_order', 'page']:
        if key in query_params:
            del query_params[key]
            
    context = {
        'page_title': 'Реєстр Актів Атестації',
        'object_list': page_obj,
        'page_obj': page_obj,
        'form': form,
        # Передаємо змінні для сортування в шаблон
        'current_sort_by': sort_by,
        'current_sort_order': sort_order,
        'sort_url_part': query_params.urlencode(),
    }
    return render(request, 'oids/lists/attestation_registered_num_list.html', context)

@login_required 
@transaction.atomic
def record_attestation_response_view(request, att_reg_sent_id=None): # Може приймати ID з URL
    attestation_registration_instance = None
    documents_to_update_qs = Document.objects.none()

    # Якщо ID передано через GET-параметр (при зміні select на сторінці)
    att_reg_sent_id_get = request.GET.get('att_reg_sent_id')
    if not att_reg_sent_id and att_reg_sent_id_get: # Якщо з URL не прийшло, але є в GET
        att_reg_sent_id = att_reg_sent_id_get
    
    if att_reg_sent_id:
        try:
            attestation_registration_instance = AttestationRegistration.objects.prefetch_related(
                'sent_documents__oid__unit', # Для відображення деталей документів у формсеті
                'sent_documents__document_type'
            ).get(
                pk=att_reg_sent_id
                # Можна додати фільтр по статусу, наприклад, тільки ті, що 'sent'
                # status=AttestationRegistrationStatusChoices.SENT 
            )
            documents_to_update_qs = attestation_registration_instance.sent_documents.all().order_by('oid__cipher')
        except AttestationRegistration.DoesNotExist:
            messages.error(request, "Обрану відправку не знайдено або на неї вже оброблена відповідь.")
            attestation_registration_instance = None # Скидаємо, щоб не показувати формсет
            documents_to_update_qs = Document.objects.none()


    if request.method == 'POST':
        # Головна форма для AttestationResponse
        # Передаємо instance, якщо він відомий, для можливого оновлення, хоча у нас OneToOne, тому це буде створення
        response_main_form = AttestationResponseMainForm(request.POST, request.FILES)
        
        # Формсет для оновлення документів
        # ID відправки має бути отримано з POST даних головної форми
        att_reg_sent_id_from_post = request.POST.get('attestation_registration_sent')
        posted_att_reg_instance_for_formset = None
        documents_for_formset_post = Document.objects.none()

        if att_reg_sent_id_from_post:
            try:
                posted_att_reg_instance_for_formset = AttestationRegistration.objects.get(pk=att_reg_sent_id_from_post)
                documents_for_formset_post = Document.objects.filter(
                    attestation_registration_sent=posted_att_reg_instance_for_formset
                ).order_by('oid__cipher')
            except AttestationRegistration.DoesNotExist:
                messages.error(request, "Помилка: відправка, зазначена у формі відповіді, не знайдена.")
        
        act_update_formset = AttestationActUpdateFormSet(request.POST, queryset=documents_for_formset_post, prefix='acts')

        if response_main_form.is_valid() and act_update_formset.is_valid():
            # Переконуємося, що attestation_registration_sent встановлено для response_main_form
            # Це має відбуватися автоматично, якщо поле є у формі і воно заповнене.
            # Якщо воно приховане і заповнене через GET, то воно має бути в request.POST
            selected_registration_for_response = posted_att_reg_instance_for_formset # Використовуємо той, що для формсету

            if selected_registration_for_response:
                # Перевіряємо, чи для цієї відправки вже є відповідь (OneToOne)
                if hasattr(selected_registration_for_response, 'response_received'):
                    messages.error(request, f"Для відправки №{selected_registration_for_response.outgoing_letter_number} вже існує відповідь.")
                else:
                    attestation_response = response_main_form.save(commit=False)
                    attestation_response.attestation_registration_sent = selected_registration_for_response
                    
                    # Тут можна встановити received_by, якщо потрібно
                    # if request.user.is_authenticated and hasattr(request.user, 'person_profile'):
                    #     attestation_response.received_by = request.user.person_profile
                    
                    attestation_response.save() 

                    # Зберігаємо зміни в документах (реєстраційні номери та дати ДССЗЗІ)
                    act_update_formset.save() 

                    # Оновлюємо статус оригінальної відправки AttestationRegistration
                    # Ця логіка тепер у AttestationResponse.save()
                    # selected_registration_for_response.status = AttestationRegistrationStatusChoices.RESPONSE_RECEIVED
                    # selected_registration_for_response.save(update_fields=['status', 'updated_at'])

                    messages.success(request, f'Відповідь ДССЗЗІ №{attestation_response.response_letter_number} успішно внесено.')
                    return redirect('oids:list_attestation_responses') 
            else:
                messages.error(request, "Не вдалося визначити відправку для збереження відповіді.")
        else:
            messages.error(request, "Будь ласка, виправте помилки у формі.")
            if not response_main_form.is_valid():
                print("Response Main Form Errors:", response_main_form.errors.as_json())
            if not act_update_formset.is_valid():
                 print("Act Update Formset Errors:", act_update_formset.errors)
                 for i, form_errors in enumerate(act_update_formset.errors):
                    if form_errors: print(f"Errors in formset form {i}: {form_errors}")

            # Якщо помилка валідації, нам потрібно зберегти attestation_registration_instance
            # щоб формсет міг бути відрендерений з правильним queryset
            if att_reg_sent_id_from_post: # Якщо ID був у POST
                 attestation_registration_instance = posted_att_reg_instance_for_formset # Використовуємо цей для контексту
            # інакше attestation_registration_instance вже встановлений з GET-параметра (якщо був)

    else: # GET request
        initial_response_data = {}
        if attestation_registration_instance:
            # Перевіряємо, чи для цієї відправки вже є відповідь (OneToOne)
            if hasattr(attestation_registration_instance, 'response_received'):
                 messages.info(request, f"Для відправки №{attestation_registration_instance.outgoing_letter_number} вже внесено відповідь. Ви можете її переглянути або редагувати (якщо реалізовано).")
                 # Можна перенаправити на сторінку редагування відповіді або просто показати повідомлення.
                 # Тут поки що просто показуємо форму з даними існуючої відповіді, якщо вона є
                 response_main_form = AttestationResponseMainForm(instance=attestation_registration_instance.response_received)
            else:
                 initial_response_data['attestation_registration_sent'] = attestation_registration_instance
                 response_main_form = AttestationResponseMainForm(initial=initial_response_data)
        else:
            response_main_form = AttestationResponseMainForm() # Порожня форма, якщо відправка не обрана

        act_update_formset = AttestationActUpdateFormSet(queryset=documents_to_update_qs, prefix='acts')

    context = {
        'response_main_form': response_main_form,
        'act_update_formset': act_update_formset,
        'page_title': 'Внести відповідь від ДССЗЗІ',
        'selected_att_reg_sent': attestation_registration_instance 
    }
    return render(request, 'oids/forms/record_attestation_response_form.html', context)

@login_required
def list_azr_registrations_view(request):
    """
    Відображає список усіх відправок АЗР на реєстрацію.
    """
    all_registrations = WorkCompletionRegistration.objects.all().order_by('-outgoing_letter_date')

    # Тут можна додати пагінацію, як на інших сторінках
    page_obj = get_paginated_page(all_registrations, request)

    context = {
        'page_title': 'Відправки АЗР на реєстрацію',
        'registrations': page_obj,
    }
    return render(request, 'oids/lists/azr_registration_list.html', context)

@login_required 
def processing_control_view(request):
    """
    Сторінка "Контроль опрацювання документів".
    Тепер тут тільки логіка для WorkRequestItem.
    """
    # today_date = datetime.date.today()
    today_date = datetime.date.today().strftime("%Y-%m-%d")
# --- ОНОВЛЕНА ЛОГІКА ---

    # 1. Створюємо підзапити, щоб "дістати" дати ОСТАННЬОГО відрядження,
    #    пов'язаного з батьківською заявкою кожного елемента.
    latest_trip_start_date = Subquery(
        Trip.objects.filter(
            work_requests=OuterRef('request')
        ).order_by('-end_date').values('start_date')[:1]
    )
    latest_trip_end_date = Subquery(
        Trip.objects.filter(
            work_requests=OuterRef('request')
        ).order_by('-end_date').values('end_date')[:1]
    )

    attestation_act_type = DocumentType.objects.filter(name__icontains="Акт атестації").first()
    ik_conclusion_type = DocumentType.objects.filter(duration_months=20).first()

    attestation_date_subquery = Document.objects.filter(
        work_request_item=OuterRef('pk'),
        document_type=attestation_act_type,
        dsszzi_registered_number__isnull=False,
        dsszzi_registered_date__isnull=False
    ).order_by('-doc_process_date').values('doc_process_date')[:1]

    ik_date_subquery = Document.objects.filter(
        work_request_item=OuterRef('pk'),
        document_type=ik_conclusion_type
    ).order_by('-doc_process_date').values('doc_process_date')[:1]

    # 2. Змінюємо головний фільтр: тепер ми шукаємо всі елементи, у яких
    #    батьківська заявка має хоча б одне відрядження.
    wri_queryset = WorkRequestItem.objects.filter(
        request__trips__isnull=False
    ).select_related(
        'request__unit', 'oid__unit'
    ).annotate(
        # 3. Додаємо дати останнього відрядження як нові поля до кожного елемента
        relevant_trip_start_date=latest_trip_start_date,
        relevant_trip_end_date=latest_trip_end_date,
        # Твої існуючі анотації
        final_attestation_date=Subquery(attestation_date_subquery),
        final_ik_date=Subquery(ik_date_subquery)
    ).distinct()
    
  
    wri_filter_form = WorkRequestItemProcessingFilterForm(request.GET or None, prefix="wri")
    if wri_filter_form.is_valid():
        if wri_filter_form.cleaned_data.get('unit'):
            wri_queryset = wri_queryset.filter(
                Q(request__unit__in=wri_filter_form.cleaned_data['unit']) | 
                Q(oid__unit__in=wri_filter_form.cleaned_data['unit'])
            ).distinct()
        if wri_filter_form.cleaned_data.get('oid'):
            wri_queryset = wri_queryset.filter(oid__in=wri_filter_form.cleaned_data['oid'])
        if wri_filter_form.cleaned_data.get('status'):
            wri_queryset = wri_queryset.filter(status=wri_filter_form.cleaned_data['status'])
        if wri_filter_form.cleaned_data.get('deadline_from'):
            wri_queryset = wri_queryset.filter(doc_processing_deadline__gte=wri_filter_form.cleaned_data['deadline_from'])
        if wri_filter_form.cleaned_data.get('deadline_to'):
            wri_queryset = wri_queryset.filter(doc_processing_deadline__lte=wri_filter_form.cleaned_data['deadline_to'])
        
         # 4. Оновлюємо фільтри по датах відрядження, щоб вони використовували нові анотовані поля
        trip_date_from = wri_filter_form.cleaned_data.get('trip_date_from') 
        if trip_date_from:
            wri_queryset = wri_queryset.filter(relevant_trip_end_date__gte=trip_date_from)
            
        trip_date_to = wri_filter_form.cleaned_data.get('trip_date_to')
        if trip_date_to:
            wri_queryset = wri_queryset.filter(relevant_trip_start_date__lte=trip_date_to)

        processed_filter = wri_filter_form.cleaned_data.get('processed')
        if processed_filter == 'yes':
            wri_queryset = wri_queryset.filter(docs_actually_processed_on__isnull=False)
        elif processed_filter == 'no':
            wri_queryset = wri_queryset.filter(docs_actually_processed_on__isnull=True)

    wri_sort_by = request.GET.get('wri_sort_by', 'doc_processing_deadline') 
    wri_sort_order = request.GET.get('wri_sort_order', 'asc')
    
    wri_valid_sorts = {
        'unit': 'request__unit__code', 
        'oid': 'oid__cipher', 
        'req_num': 'request__incoming_number', 
        # ... (решта полів)
        'status': 'status',
        # 5. Вказуємо, що сортування по 'trip_dates' має використовувати наше нове анотоване поле
        'trip_dates': 'relevant_trip_end_date', 
        'deadline': 'doc_processing_deadline', 
        'proc_date': 'docs_actually_processed_on'
    }
    wri_order_by_field_key = wri_sort_by.lstrip('-')
    # Отримуємо фактичне ім'я поля для сортування з wri_valid_sorts
    # Якщо ключ 'trip_dates', то order_by_field_for_db буде 'relevant_trip_end_date'
    order_by_field_for_db = wri_valid_sorts.get(wri_order_by_field_key, 'doc_processing_deadline')
    
    current_wri_sort_order_for_template = wri_sort_order # Для передачі в шаблон
    
    # Визначаємо напрямок сортування
    if wri_sort_by.startswith('-'):
        if wri_sort_order == 'asc': # Користувач клікнув для зміни на asc
            final_wri_order_by_field = order_by_field_for_db # Прибираємо мінус
            current_wri_sort_order_for_template = 'asc'
        else: # Залишаємо desc (напрямок вже є у wri_sort_by)
            final_wri_order_by_field = f"-{order_by_field_for_db}" # Додаємо мінус, якщо його не було
            current_wri_sort_order_for_template = 'desc'
    else: # Якщо wri_sort_by не починається з '-'
        if wri_sort_order == 'desc':
            final_wri_order_by_field = f"-{order_by_field_for_db}"
            current_wri_sort_order_for_template = 'desc'
        else: # asc
            final_wri_order_by_field = order_by_field_for_db
            current_wri_sort_order_for_template = 'asc'
            
    if final_wri_order_by_field:
        print(f"DEBUG WRI Sorting by: {final_wri_order_by_field}")
        wri_queryset = wri_queryset.order_by(final_wri_order_by_field, 'request__incoming_date', 'oid__cipher')
    else: # За замовчуванням
        wri_queryset = wri_queryset.order_by('doc_processing_deadline', 'request__incoming_date', 'oid__cipher')
        current_wri_sort_order_for_template = 'asc' 
    page_obj = get_paginated_page(wri_queryset, request)

    all_units = Unit.objects.all().order_by('code')

    context = {
        'page_title': 'Контроль опрацювання',
        'wri_filter_form': wri_filter_form,
        'work_request_items': page_obj, 
        'all_units': all_units, 
        'today_date': today_date,
        'current_wri_sort_by': wri_sort_by.lstrip('-'), # Чисте ім'я поля для порівняння в шаблоні
        'current_wri_sort_order': current_wri_sort_order_for_template, # Напрямок для наступного кліку
        'ReviewResultChoices': DocumentReviewResultChoices,
    }
    return render(request, 'oids/processing_control_dashboard.html', context)

@login_required
def technical_task_control_view(request):
    """
    Нова view для сторінки "Контроль опрацювання ТЗ/МЗ".
    """
    today_date = datetime.date.today()

    tt_queryset = TechnicalTask.objects.select_related('oid__unit', 'reviewed_by')

    tt_filter_form = TechnicalTaskFilterForm(request.GET or None, prefix="tt")
    if tt_filter_form.is_valid():
        if tt_filter_form.cleaned_data.get('unit'):
            tt_queryset = tt_queryset.filter(oid__unit__in=tt_filter_form.cleaned_data['unit'])
        if tt_filter_form.cleaned_data.get('oid'):
            tt_queryset = tt_queryset.filter(oid__in=tt_filter_form.cleaned_data['oid'])
        if tt_filter_form.cleaned_data.get('review_result'):
            tt_queryset = tt_queryset.filter(review_result=tt_filter_form.cleaned_data['review_result'])

    # Сортування для ТЗ
    tt_sort_by = request.GET.get('tt_sort_by', '-read_till_date')
    tt_sort_order = request.GET.get('tt_sort_order', 'asc')
    tt_valid_sorts = {
        'unit': 'oid__unit__code', 'oid': 'oid__cipher', 'input_num': 'input_number',
        'input_date': 'input_date', 'deadline': 'read_till_date', 'status': 'review_result',
        'executor': 'reviewed_by__full_name', 'exec_date': 'updated_at'
    }
    tt_order_by_field_key = tt_sort_by.lstrip('-')
    tt_order_by_field = tt_valid_sorts.get(tt_order_by_field_key, '-read_till_date')

    if tt_sort_by.startswith('-'):
        if tt_sort_order == 'asc':
            tt_order_by_field = tt_order_by_field_key
    elif tt_sort_order == 'desc':
        tt_order_by_field = f"-{tt_order_by_field}"

    tt_queryset = tt_queryset.order_by(tt_order_by_field, 'input_number')
    
    page_obj = get_paginated_page(tt_queryset, request)

    # Готуємо контекст для передачі в шаблон
    context = {
        'page_title': 'Контроль опрацювання ТЗ/МЗ',
        'tt_filter_form': tt_filter_form,
        'technical_tasks': page_obj,
        'today_date': today_date,
        'current_tt_sort_by': tt_sort_by.lstrip('-'),
        'current_tt_sort_order': 'desc' if tt_sort_by.startswith('-') else tt_sort_order,
        'ReviewResultChoices': DocumentReviewResultChoices,
    }
    return render(request, 'oids/technical_task_control.html', context)

@transaction.atomic
def start_declaration_process_view(request):
    template_name = "ДСК-Декларація"
    try:
        process_template = ProcessTemplate.objects.get(name=template_name, is_active=True)
    except ProcessTemplate.DoesNotExist:
        messages.error(request, f"Шаблон процесу '{template_name}' не знайдено або не активний.")
        return redirect('oids:dashboard') # або на головну

    if request.method == 'POST':
        form = DeclarationProcessStartForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            
            # 1. Створення ОІД
            new_oid = OID.objects.create(
                unit=data['unit'],
                cipher=data['cipher'],
                oid_type=data['oid_type'],
                sec_level=data['sec_level'],
                status=data['initial_oid_status'],
                room=data['room'],
                full_name=data.get('full_name', ''),
                pemin_sub_type=data['pemin_sub_type'],
                serial_number=data.get('serial_number', ''),
                inventory_number=data.get('inventory_number', ''),
                note=data.get('note', '')
            )
            messages.success(request, f"ОІД {new_oid.cipher} успішно створено.")

            # 2. Створення початкового документа "Декларація"
            declaration_doc_type = DocumentType.objects.get(name="Декларація відповідності")
            declaration_doc = Document.objects.create(
                oid=new_oid,
                document_type=declaration_doc_type,
                # Можна встановити автора, якщо є логіка визначення поточного користувача
                # author=request.user.person,
                document_number=f"ДЕКЛ-{new_oid.cipher}", # Генерація унікального номера
                doc_process_date=datetime.date.today(),
                work_date=datetime.date.today(),
                processing_status=DocumentProcessingStatusChoices.DRAFT # Початковий статус
            )
            messages.info(request, f"Створено чернетку документа '{declaration_doc.document_number}'.")

            # 3. Запуск процесу для ОІД
            oid_process = OIDProcess.objects.create(oid=new_oid, template=process_template)
            
            # 4. Створення екземплярів кроків
            for step_template in process_template.steps.all():
                OIDProcessStepInstance.objects.create(
                    oid_process=oid_process,
                    process_step=step_template
                )
            
            messages.success(request, f"Для ОІД {new_oid.cipher} запущено процес '{process_template.name}'.")
            return redirect('oids:oid_detail', pk=new_oid.pk) # Перенаправляємо на сторінку ОІД
            
    else:
        form = DeclarationProcessStartForm()

    context = {
        'form': form,
        'page_title': f'Ініціація процесу: {template_name}'
    }
    return render(request, 'oids/generic_form.html', context)

@login_required
def azr_documents_list_view(request):
    """
    Відображає список всіх документів типу "АЗР" з фільтрацією та сортуванням.
    Також обробляє експорт в Excel.
    """
    try:
        azr_doc_type = DocumentType.objects.get(name__icontains="Акт завершення")
    except DocumentType.DoesNotExist:
        messages.error(request, "Тип документу 'Акт завершення' не знайдено.")
        return redirect('oids:main_dashboard')

    queryset = Document.objects.filter(document_type=azr_doc_type).select_related('oid__unit')
    
    # --- Фільтрація (залишається без змін) ---
    filter_form = AzrDocumentFilterForm(request.GET or None)
    if filter_form.is_valid():
        selected_units = filter_form.cleaned_data.get('unit')
        if selected_units:
            # Використовуємо __in для фільтрації по списку значень
            queryset = queryset.filter(oid__unit__in=selected_units)

        selected_oids = filter_form.cleaned_data.get('oid')
        if selected_oids:
            # Використовуємо __in для фільтрації по списку значень
            queryset = queryset.filter(oid__in=selected_oids)
            
		# if filter_form.cleaned_data.get('unit'):
        #     queryset = queryset.filter(oid__unit__in=filter_form.cleaned_data['unit'])
        # if filter_form.cleaned_data.get('oid'):
        #     queryset = queryset.filter(oid__in=filter_form.cleaned_data['oid'])
        # if filter_form.cleaned_data.get('document_number'):
        #     queryset = queryset.filter(document_number__icontains=filter_form.cleaned_data['document_number'])
        # if filter_form.cleaned_data.get('registered_number'):
        #     queryset = queryset.filter(dsszzi_registered_number__icontains=filter_form.cleaned_data['registered_number'])
        # if filter_form.cleaned_data.get('date_from'):
        #     queryset = queryset.filter(doc_process_date__gte=filter_form.cleaned_data['date_from'])
        # if filter_form.cleaned_data.get('date_to'):
        #     queryset = queryset.filter(doc_process_date__lte=filter_form.cleaned_data['date_to'])
 
        pass

    # --- Сортування (залишається без змін) ---
    sort_by = request.GET.get('sort', '-doc_process_date')
    valid_sort_fields = ['oid__unit__code', 'oid__cipher', 'document_number', 'doc_process_date', 'dsszzi_registered_number', 'dsszzi_registered_date']
    if sort_by.lstrip('-') in valid_sort_fields:
        queryset = queryset.order_by(sort_by)

    # --- Експорт в Excel ---
    if 'export' in request.GET:
        # Визначаємо, які поля та з якими назвами ми хочемо експортувати
        columns = {
            'oid__unit__code': 'ВЧ',
            'oid__cipher': 'Шифр ОІД',
            'document_number': 'Підг. №',
            'doc_process_date': 'Від',
            'dsszzi_registered_number': 'Зареєстрований № АЗР',
            'dsszzi_registered_date': 'Від якого числа',
        }
        # Викликаємо нашу універсальну функцію
        return export_to_excel(
            queryset, 
            columns, 
            filename='azr_documents_export.xlsx'
        )
    
    # --- Пагінація (залишається без змін) ---
    page_obj = get_paginated_page(queryset, request)
    
    # --- НОВА ЛОГІКА: Підготовка URL для сортування ---
    query_params = request.GET.copy()
    # Видаляємо параметри 'sort' та 'page', щоб уникнути їх дублювання в URL
    if 'sort' in query_params:
        del query_params['sort']
    if 'page' in query_params:
        del query_params['page']
    # --------------------------------------------------

    context = {
        'page_title': 'Реєстр Актів Завершення Робіт (АЗР)',
        'documents': page_obj,
        'filter_form': filter_form,
        'current_sort': sort_by,
        'sort_url_part': query_params.urlencode(), # <-- Передаємо готову частину URL в шаблон
    }
    return render(request, 'oids/lists/azr_document_list.html', context)


